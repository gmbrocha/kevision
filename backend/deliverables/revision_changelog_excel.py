"""Revision changelog Excel exporter.

This module owns the current workbook layout used for the review/export
deliverable. The schema is documented in `docs/revision_changelog_format.md`.

One row per (sheet, detail) group of approved change items. Sub-items inside a
group stack as numbered bullets (`1)`, `2)`, ...) inside the merged
`Scope Included` cell. The cropped cloud image is embedded in column F.

Columns intentionally preserve the current downstream workbook shape,
including the historical `Qoute Received?` typo.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from ..crop_adjustments import build_selected_review_overlay_image, selected_review_page_boxes
from ..review import change_item_needs_attention
from ..review_queue import ordered_change_items, visible_change_items
from ..revision_state.models import ChangeItem, CloudCandidate, RevisionSet, SheetVersion
from ..utils import clean_display_text
from ..workspace import WorkspaceStore
from .crop_comparison import build_cloud_comparison_image, find_previous_sheet_version


COLUMNS: list[tuple[str, float]] = [
    ("Correlation ", 12),
    ("Drawing #", 12),
    ("Revision # ", 24),
    ("Detail #", 18),
    ("Scope Included ", 60),
    ("Detail View ", 40),
    ("", 30),
    ("Responsible Contractor", 22),
    ("Cost?", 10),
    ("Qoute Received?", 16),
]

ROWS_PER_GROUP = 14
SCOPE_COL = 5
DETAIL_VIEW_COL = 6
CROP_TARGET_PX = (520, ROWS_PER_GROUP * 18)
HEADER_FILL = "111827"
HEADER_TEXT = "FFFFFF"
BLOCK_FILL = "FFFFFF"
BLOCK_ALT_FILL = "F8FAFC"
META_FILL = "EEF3F8"
META_ALT_FILL = "E7EEF6"
CROP_FILL = "FFFFFF"
BORDER_COLOR = "CDD6E2"
BORDER_STRONG_COLOR = "7D8DA1"
ACCENT_COLOR = "0F766E"


@dataclass
class RevisionChangelogRow:
    correlation: str
    drawing: str
    revision: str
    detail: str
    scope_lines: list[str]
    crop_path: str | None
    cloud_id: str | None
    sheet_version_id: str | None
    item_ids: list[str]


def write_revision_changelog(store: WorkspaceStore, output_path: Path) -> Path:
    rows = _build_rows(store)
    _write_workbook(store, rows, output_path)
    return output_path


def _build_rows(store: WorkspaceStore) -> list[RevisionChangelogRow]:
    revision_sets_by_id = {rs.id: rs for rs in store.data.revision_sets}
    sheets_by_id = {sheet.id: sheet for sheet in store.data.sheets}
    clouds_by_id = {cloud.id: cloud for cloud in store.data.clouds}

    approved = [item for item in visible_change_items(store.data.change_items) if item.status == "approved"]
    groups: dict[tuple[str, str, str], list[ChangeItem]] = {}
    for item in approved:
        key = _group_key(item)
        groups.setdefault(key, []).append(item)

    grouped_rows: list[tuple[str, int, RevisionChangelogRow]] = []
    sheet_counters: dict[str, int] = {}
    for (sheet_id, group_kind, group_ref), items in groups.items():
        sheet_counters[sheet_id] = sheet_counters.get(sheet_id, 0) + 1
        seq = sheet_counters[sheet_id]
        canonical_sheet = items[0]
        sheet = sheets_by_id.get(canonical_sheet.sheet_version_id)
        revision_set = revision_sets_by_id.get(sheet.revision_set_id) if sheet else None
        cloud = _pick_cloud(items, clouds_by_id)
        crop_path = _crop_path(cloud)
        detail_ref = group_ref if group_kind == "detail" else None
        revision_row = RevisionChangelogRow(
            correlation=_format_correlation(sheet_id, seq),
            drawing=_format_drawing(sheet_id),
            revision=_format_revision(revision_set),
            detail=_format_detail(detail_ref),
            scope_lines=_collect_scope_lines(items),
            crop_path=crop_path,
            cloud_id=cloud.id if cloud else None,
            sheet_version_id=sheet.id if sheet else None,
            item_ids=[item.id for item in items],
        )
        grouped_rows.append((sheet_id, seq, revision_row))

    grouped_rows.sort(key=lambda triple: (triple[0], triple[1]))
    return [row for _, _, row in grouped_rows]


def _group_key(item: ChangeItem) -> tuple[str, str, str]:
    if item.detail_ref:
        return (item.sheet_id, "detail", item.detail_ref)
    if item.cloud_candidate_id:
        return (item.sheet_id, "cloud", item.cloud_candidate_id)
    return (item.sheet_id, "item", item.id)


def _format_correlation(sheet_id: str, seq: int) -> str:
    digits = "".join(re.findall(r"\d+", sheet_id))
    if not digits:
        return f"{sheet_id}.{seq}"
    return f"{digits}.{seq}"


def _format_drawing(sheet_id: str) -> str:
    match = re.match(r"^([A-Za-z]+)[-\s]?(\d.*)$", sheet_id.strip())
    if not match:
        return sheet_id
    letters, rest = match.groups()
    return f"{letters.upper()}-{rest}"


def _format_revision(revision_set: RevisionSet | None) -> str:
    if revision_set is None:
        return ""
    label = clean_display_text(revision_set.label) or f"Revision #{revision_set.set_number}"
    base = re.split(r"\s+-\s+", label, maxsplit=1)[0]
    if revision_set.set_date:
        return f"{base} - {revision_set.set_date}"
    return base


def _format_detail(detail_ref: str | None) -> str:
    if not detail_ref:
        return "N/A - Cloud Only "
    cleaned = clean_display_text(detail_ref)
    if not cleaned:
        return "N/A - Cloud Only "
    if cleaned.lower().startswith("detail"):
        return f"{cleaned} "
    return f"Detail {cleaned} "


def _collect_scope_lines(items: list[ChangeItem]) -> list[str]:
    seen: set[str] = set()
    lines: list[str] = []
    for item in items:
        text = clean_display_text(item.reviewer_text or item.raw_text)
        if not text:
            continue
        for piece in re.split(r"\s*\|\s*|\s*;\s*|[\r\n]+", text):
            line = clean_display_text(piece).strip(" -|:,.;")
            if not line:
                continue
            normalized = line.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            lines.append(line)
    return lines


def _pick_cloud(items: list[ChangeItem], clouds_by_id: dict[str, CloudCandidate]) -> CloudCandidate | None:
    for item in items:
        cloud_id = item.cloud_candidate_id
        if not cloud_id:
            continue
        cloud = clouds_by_id.get(cloud_id)
        if cloud is not None:
            return cloud
    return None


def _crop_path(cloud: CloudCandidate | None) -> str | None:
    if cloud is None:
        return None
    path = cloud.image_path
    if path and Path(path).exists():
        return path
    return None


def _write_workbook(store: WorkspaceStore, rows: list[RevisionChangelogRow], output_path: Path) -> None:
    revision_sets_by_id = {rs.id: rs for rs in store.data.revision_sets}
    sheets_by_id = {sheet.id: sheet for sheet in store.data.sheets}
    clouds_by_id = {cloud.id: cloud for cloud in store.data.clouds}
    items_by_id = {item.id: item for item in visible_change_items(store.data.change_items)}
    comparison_dir = output_path.parent / f"{output_path.stem}_comparison_images"

    wb = Workbook()
    wb.properties.creator = "ScopeLedger"
    wb.properties.title = "Revision Changelog"
    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_summary_sheet(summary_ws, store, rows)

    ws = wb.create_sheet("Sheet1")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = ACCENT_COLOR

    header_font = Font(bold=True, color=HEADER_TEXT)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    thin = Side(border_style="thin", color=BORDER_COLOR)
    strong = Side(border_style="medium", color=BORDER_STRONG_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 30

    cursor = 2
    for row_index, revision_row in enumerate(rows):
        block_top = cursor
        block_bottom = cursor + ROWS_PER_GROUP - 1
        is_alt = row_index % 2 == 1

        ws.cell(row=block_top, column=1, value=revision_row.correlation).alignment = wrap
        ws.cell(row=block_top, column=2, value=revision_row.drawing).alignment = wrap
        ws.cell(row=block_top, column=3, value=revision_row.revision).alignment = wrap
        ws.cell(row=block_top, column=4, value=revision_row.detail).alignment = wrap

        scope_text = _format_scope_text(revision_row.scope_lines)
        scope_cell = ws.cell(row=block_top, column=SCOPE_COL, value=scope_text)
        scope_cell.alignment = wrap

        block_fill = PatternFill("solid", fgColor=BLOCK_ALT_FILL if is_alt else BLOCK_FILL)
        meta_fill = PatternFill("solid", fgColor=META_ALT_FILL if is_alt else META_FILL)
        crop_fill = PatternFill("solid", fgColor=CROP_FILL)
        for r in range(block_top, block_bottom + 1):
            ws.row_dimensions[r].height = 19
            for c in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = meta_fill if c <= 4 else (crop_fill if c == DETAIL_VIEW_COL else block_fill)
                cell.alignment = center_wrap if c <= 4 else wrap
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=strong if r == block_top else thin,
                    bottom=strong if r == block_bottom else thin,
                )
                if r == block_top and c <= 4:
                    cell.font = Font(bold=True, color="111827")

        if block_bottom > block_top:
            ws.merge_cells(
                start_row=block_top, start_column=SCOPE_COL,
                end_row=block_bottom, end_column=SCOPE_COL,
            )
            ws.merge_cells(
                start_row=block_top, start_column=DETAIL_VIEW_COL,
                end_row=block_bottom, end_column=DETAIL_VIEW_COL,
            )

        image_path = _detail_view_image_path(
            store,
            revision_row,
            sheets_by_id=sheets_by_id,
            clouds_by_id=clouds_by_id,
            items_by_id=items_by_id,
            revision_sets_by_id=revision_sets_by_id,
            comparison_dir=comparison_dir,
            row_index=row_index,
        )
        if image_path:
            try:
                _embed_image(ws, image_path, anchor_row=block_top, anchor_col=DETAIL_VIEW_COL)
            except Exception:
                ws.cell(row=block_top, column=DETAIL_VIEW_COL, value=f"[crop: {Path(image_path).name}]")

        cursor = block_bottom + 1

    last_column = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_column}{max(ws.max_row, 1)}"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    review_ws = wb.create_sheet("Review Flags")
    _write_review_flags_sheet(review_ws, store, revision_sets_by_id, sheets_by_id)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_summary_sheet(ws, store: WorkspaceStore, rows: list[RevisionChangelogRow]) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = HEADER_FILL
    for column, width in {
        "A": 4,
        "B": 24,
        "C": 18,
        "D": 24,
        "E": 18,
        "F": 38,
        "G": 22,
    }.items():
        ws.column_dimensions[column].width = width

    title_fill = PatternFill("solid", fgColor=HEADER_FILL)
    accent_fill = PatternFill("solid", fgColor=ACCENT_COLOR)
    panel_fill = PatternFill("solid", fgColor="F8FAFC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(border_style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, color=HEADER_TEXT, size=20)
    subtitle_font = Font(color="D1D5DB", size=11)
    heading_font = Font(bold=True, color="111827", size=13)
    label_font = Font(bold=True, color="475569", size=10)
    metric_font = Font(bold=True, color="111827", size=24)
    body_font = Font(color="334155", size=11)
    note_font = Font(color="475569", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A1:G3")
    title = ws["A1"]
    title.value = "ScopeLedger Revision Review"
    title.fill = title_fill
    title.font = title_font
    title.alignment = Alignment(horizontal="left", vertical="center")
    for row in range(1, 4):
        ws.row_dimensions[row].height = 24
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = title_fill

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells("A4:G4")
    subtitle = ws["A4"]
    subtitle.value = f"Generated {generated_at} | Local review workbook"
    subtitle.fill = title_fill
    subtitle.font = subtitle_font
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    change_items = visible_change_items(store.data.change_items)
    approved = [item for item in change_items if item.status == "approved"]
    pending = [item for item in change_items if item.status == "pending"]
    rejected = [item for item in change_items if item.status == "rejected"]
    active_sheets = [sheet for sheet in store.data.sheets if sheet.status == "active"]
    superseded_sheets = [sheet for sheet in store.data.sheets if sheet.status == "superseded"]
    crop_rows = [row for row in rows if row.crop_path or row.cloud_id]
    cloud_rows = [
        item
        for item in approved
        if item.provenance.get("source") == "visual-region"
        and item.provenance.get("extraction_method") == "cloudhammer_manifest"
    ]

    ws.merge_cells("A6:G6")
    section = ws["A6"]
    section.value = "Package Summary"
    section.font = heading_font
    section.alignment = left

    cards = [
        ("Revision Sets", len(store.data.revision_sets)),
        ("Current Sheets", len(active_sheets)),
        ("Superseded Sheets", len(superseded_sheets)),
        ("Accepted Changes", len(approved)),
        ("Workbook Rows", len(rows)),
        ("Crop Images", len(crop_rows)),
        ("Needs Review", len(pending)),
        ("Rejected", len(rejected)),
    ]
    card_positions = [("B", "C", 8), ("D", "E", 8), ("F", "G", 8), ("B", "C", 12), ("D", "E", 12), ("F", "G", 12), ("B", "C", 16), ("D", "E", 16)]
    for (label, value), (start_col, end_col, row) in zip(cards, card_positions):
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
        ws.merge_cells(f"{start_col}{row + 1}:{end_col}{row + 2}")
        label_cell = ws[f"{start_col}{row}"]
        value_cell = ws[f"{start_col}{row + 1}"]
        label_cell.value = label
        label_cell.fill = panel_fill
        label_cell.font = label_font
        label_cell.alignment = center
        value_cell.value = value
        value_cell.fill = white_fill
        value_cell.font = metric_font
        value_cell.alignment = center
        for r in range(row, row + 3):
            for c in range(ws[start_col + str(r)].column, ws[end_col + str(r)].column + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if cell.fill.fill_type is None:
                    cell.fill = white_fill

    ws.merge_cells("A20:G20")
    workflow_heading = ws["A20"]
    workflow_heading.value = "How to Use This Workbook"
    workflow_heading.font = heading_font
    workflow_heading.alignment = left

    guidance = [
        "Start on Sheet1 for the detailed change list and embedded crop evidence.",
        "Use the crop image as the source of truth while scope text extraction is still being built.",
        "Contractor, cost, and quote fields are intentionally left blank for downstream review.",
        "Source PDFs and sensitive project data remain local unless a separate security approval allows otherwise.",
    ]
    for offset, text in enumerate(guidance, start=21):
        ws.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=7)
        ws.cell(row=offset, column=1, value=offset - 20).fill = accent_fill
        ws.cell(row=offset, column=1).font = Font(bold=True, color=HEADER_TEXT)
        ws.cell(row=offset, column=1).alignment = center
        cell = ws.cell(row=offset, column=2, value=text)
        cell.font = body_font
        cell.alignment = left

    ws.merge_cells("A27:G29")
    note = ws["A27"]
    note.value = (
        "Current MVP status: cloud/crop evidence is available for review. "
        "Legend parsing, keynote interpretation, detail-reference extraction, "
        "and polished scope descriptions are planned next steps."
    )
    note.fill = panel_fill
    note.font = note_font
    note.alignment = left
    for row in range(27, 30):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = panel_fill
            ws.cell(row=row, column=col).border = border

    ws.merge_cells("A31:G31")
    footer = ws["A31"]
    footer.value = f"Detected-region accepted rows: {len(cloud_rows)}"
    footer.font = Font(bold=True, color=ACCENT_COLOR)
    footer.alignment = left


def _format_scope_text(lines: list[str]) -> str:
    if not lines:
        return ""
    if len(lines) == 1:
        return lines[0]
    return "\n".join(f"{i}) {line}" for i, line in enumerate(lines, 1))


def _write_review_flags_sheet(
    ws,
    store: WorkspaceStore,
    revision_sets_by_id: dict[str, RevisionSet],
    sheets_by_id: dict[str, SheetVersion],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "F59E0B"
    headers = [
        ("Change ID", 18),
        ("Status", 12),
        ("Needs Review", 14),
        ("Review Reason", 34),
        ("Drawing #", 14),
        ("Detail #", 18),
        ("Revision #", 24),
        ("Source PDF", 42),
        ("Page", 8),
        ("Source Kind", 16),
        ("Extraction Method", 18),
        ("Extraction Reason", 22),
        ("Extraction Signal", 16),
        ("Reviewer Notes", 34),
        ("Scope Text", 72),
    ]
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_TEXT)
    thin = Side(border_style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    for column, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[1].height = 32

    rows = [
        item
        for item in visible_change_items(store.data.change_items)
        if item.status != "rejected"
    ]
    rows = ordered_change_items(rows)
    for row_index, item in enumerate(rows, 2):
        sheet = sheets_by_id.get(item.sheet_version_id)
        revision_set = revision_sets_by_id.get(sheet.revision_set_id) if sheet else None
        needs_review = item.status == "pending" and change_item_needs_attention(item)
        values = [
            item.id,
            item.status,
            "Yes" if needs_review else "No",
            _review_reason(item),
            _format_drawing(item.sheet_id),
            _format_detail(item.detail_ref),
            _format_revision(revision_set),
            store.display_path(sheet.source_pdf) if sheet else "",
            sheet.page_number if sheet else "",
            item.provenance.get("source", ""),
            item.provenance.get("extraction_method", ""),
            item.provenance.get("scope_text_reason", ""),
            item.provenance.get("extraction_signal", ""),
            clean_display_text(item.reviewer_notes),
            clean_display_text(item.reviewer_text or item.raw_text),
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            cell.border = border
            if column == 3 and value == "Yes":
                cell.font = Font(bold=True, color="92400E")
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
        ws.row_dimensions[row_index].height = 34

    last_column = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_column}{max(ws.max_row, 1)}"


def _review_reason(item: ChangeItem) -> str:
    provenance = item.provenance or {}
    reasons: list[str] = []
    scope_reason = str(provenance.get("scope_text_reason") or "")
    scope_labels = {
        "text-layer-near-cloud": "PDF text found near cloud",
        "ocr-near-cloud": "OCR text found near cloud - verify",
        "no-readable-text": "No readable scope text near cloud",
        "leader-or-callout-only": "Leader/detail callout only",
        "needs-reviewer-rewrite": "Nearby text needs reviewer rewrite",
        "index-or-title-noise": "Nearby text appears to be index/title-block noise",
    }
    if scope_reason:
        reasons.append(scope_labels.get(scope_reason, scope_reason))
    if provenance.get("source") == "visual-region" and not item.detail_ref:
        reasons.append("No detail reference captured")
    try:
        signal = float(provenance.get("extraction_signal", 1.0))
    except (TypeError, ValueError):
        signal = 1.0
    if signal < 0.48:
        reasons.append("Low extraction signal")
    if not reasons and provenance.get("source") == "narrative":
        reasons.append("Narrative-derived item")
    return "; ".join(dict.fromkeys(reasons))


def _detail_view_image_path(
    store: WorkspaceStore,
    revision_row: RevisionChangelogRow,
    *,
    sheets_by_id: dict[str, SheetVersion],
    clouds_by_id: dict[str, CloudCandidate],
    items_by_id: dict[str, ChangeItem],
    revision_sets_by_id: dict[str, RevisionSet],
    comparison_dir: Path,
    row_index: int,
) -> str | None:
    cloud = clouds_by_id.get(revision_row.cloud_id or "")
    sheet = sheets_by_id.get(revision_row.sheet_version_id or "")
    if cloud is None or sheet is None:
        return revision_row.crop_path
    selected_item = next((items_by_id[item_id] for item_id in revision_row.item_ids if item_id in items_by_id), None)
    if selected_item:
        selected_overlay = comparison_dir / f"{row_index + 1:04d}_{cloud.id}_selected.png"
        generated_overlay = build_selected_review_overlay_image(
            store,
            selected_item,
            cloud,
            selected_overlay,
            include_all=False,
        )
        if generated_overlay:
            return str(generated_overlay)

    previous_sheet = find_previous_sheet_version(sheet, list(sheets_by_id.values()), revision_sets_by_id)
    comparison_path = comparison_dir / f"{row_index + 1:04d}_{cloud.id}.png"
    selected_boxes = selected_review_page_boxes(selected_item, cloud) if selected_item else []
    generated = build_cloud_comparison_image(
        store,
        cloud=cloud,
        current_sheet=sheet,
        previous_sheet=previous_sheet,
        output_path=comparison_path,
        highlight_bboxes=selected_boxes or None,
    )
    if generated:
        return str(generated)
    return revision_row.crop_path


def _embed_image(ws, image_path: str, *, anchor_row: int, anchor_col: int) -> None:
    """Resize the source crop to fit the merged cell area and anchor it there.

    openpyxl's image insertion takes pixel dimensions; we resize through Pillow
    first so the embedded PNG is small (Kevin's file uses ~500px-wide crops).
    """
    target_w, target_h = CROP_TARGET_PX
    with PILImage.open(image_path) as src:
        src.load()
        src = src.convert("RGB") if src.mode in ("RGBA", "P") else src
        src.thumbnail((target_w, target_h), PILImage.LANCZOS)
        buf = io.BytesIO()
        src.save(buf, format="PNG", optimize=True)
        buf.seek(0)
    img = XLImage(buf)
    img.anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
    ws.add_image(img)
