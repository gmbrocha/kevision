"""Kevin-shaped changelog Excel exporter.

Reverse-engineered from `mod_5_changelog.xlsx` (Kevin's in-progress template,
4/21/26). Schema is documented in `docs/kevin_changelog_format.md`. Keep this
module the single owner of that layout so future column tweaks are localized.

One row per (sheet, detail) group of approved change items. Sub-items inside a
group stack as numbered bullets (`1)`, `2)`, ...) inside the merged
`Scope Included` cell. The cropped cloud image is embedded in column F.

Columns mirror Kevin's file exactly, including the `Qoute Received?` typo.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .models import ChangeItem, RevisionSet, SheetVersion
from .utils import clean_display_text
from .workspace import WorkspaceStore


COLUMNS: list[tuple[str, float]] = [
    ("Correlation ", 12),
    ("Drawing #", 12),
    ("Revision # ", 24),
    ("Detail #", 18),
    ("Scope Included ", 60),
    ("Detail View ", 40),
    ("", 30),
    ("Responsible Contractor", 22),
    ("Cost?", 10),
    ("Qoute Received?", 16),
]

ROWS_PER_GROUP = 14
SCOPE_COL = 5
DETAIL_VIEW_COL = 6
CROP_TARGET_PX = (520, ROWS_PER_GROUP * 18)


@dataclass
class KevinRow:
    correlation: str
    drawing: str
    revision: str
    detail: str
    scope_lines: list[str]
    crop_path: str | None
    item_ids: list[str]


def write_kevin_changelog(store: WorkspaceStore, output_path: Path) -> Path:
    rows = _build_rows(store)
    _write_workbook(rows, output_path)
    return output_path


def _build_rows(store: WorkspaceStore) -> list[KevinRow]:
    revision_sets_by_id = {rs.id: rs for rs in store.data.revision_sets}
    sheets_by_id = {sheet.id: sheet for sheet in store.data.sheets}
    clouds_by_id = {cloud.id: cloud for cloud in store.data.clouds}

    approved = [item for item in store.data.change_items if item.status == "approved"]
    groups: dict[tuple[str, str | None], list[ChangeItem]] = {}
    for item in approved:
        key = (item.sheet_id, item.detail_ref or None)
        groups.setdefault(key, []).append(item)

    grouped_rows: list[tuple[str, int, KevinRow]] = []
    sheet_counters: dict[str, int] = {}
    for (sheet_id, detail_ref), items in groups.items():
        sheet_counters[sheet_id] = sheet_counters.get(sheet_id, 0) + 1
        seq = sheet_counters[sheet_id]
        canonical_sheet = items[0]
        sheet = sheets_by_id.get(canonical_sheet.sheet_version_id)
        revision_set = revision_sets_by_id.get(sheet.revision_set_id) if sheet else None
        crop_path = _pick_crop_path(items, clouds_by_id)
        kevin_row = KevinRow(
            correlation=_format_correlation(sheet_id, seq),
            drawing=_format_drawing(sheet_id),
            revision=_format_revision(revision_set),
            detail=_format_detail(detail_ref),
            scope_lines=_collect_scope_lines(items),
            crop_path=crop_path,
            item_ids=[item.id for item in items],
        )
        grouped_rows.append((sheet_id, seq, kevin_row))

    grouped_rows.sort(key=lambda triple: (triple[0], triple[1]))
    return [row for _, _, row in grouped_rows]


def _format_correlation(sheet_id: str, seq: int) -> str:
    digits = "".join(re.findall(r"\d+", sheet_id))
    if not digits:
        return f"{sheet_id}.{seq}"
    return f"{digits}.{seq}"


def _format_drawing(sheet_id: str) -> str:
    match = re.match(r"^([A-Za-z]+)[-\s]?(\d.*)$", sheet_id.strip())
    if not match:
        return sheet_id
    letters, rest = match.groups()
    return f"{letters.upper()}-{rest}"


def _format_revision(revision_set: RevisionSet | None) -> str:
    if revision_set is None:
        return ""
    label = clean_display_text(revision_set.label) or f"Revision #{revision_set.set_number}"
    base = re.split(r"\s+-\s+", label, maxsplit=1)[0]
    if revision_set.set_date:
        return f"{base} - {revision_set.set_date}"
    return base


def _format_detail(detail_ref: str | None) -> str:
    if not detail_ref:
        return "N/A - Cloud Only "
    cleaned = clean_display_text(detail_ref)
    if not cleaned:
        return "N/A - Cloud Only "
    if cleaned.lower().startswith("detail"):
        return f"{cleaned} "
    return f"Detail {cleaned} "


def _collect_scope_lines(items: list[ChangeItem]) -> list[str]:
    seen: set[str] = set()
    lines: list[str] = []
    for item in items:
        text = clean_display_text(item.reviewer_text or item.raw_text)
        if not text:
            continue
        for piece in re.split(r"\s*\|\s*|\s*;\s*|[\r\n]+", text):
            line = clean_display_text(piece).strip(" -|:,.;")
            if not line:
                continue
            normalized = line.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            lines.append(line)
    return lines


def _pick_crop_path(items: list[ChangeItem], clouds_by_id: dict[str, object]) -> str | None:
    for item in items:
        cloud_id = item.cloud_candidate_id
        if not cloud_id:
            continue
        cloud = clouds_by_id.get(cloud_id)
        if cloud is None:
            continue
        path = getattr(cloud, "image_path", "")
        if path and Path(path).exists():
            return path
    return None


def _write_workbook(rows: list[KevinRow], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22

    cursor = 2
    for kevin_row in rows:
        block_top = cursor
        block_bottom = cursor + ROWS_PER_GROUP - 1

        ws.cell(row=block_top, column=1, value=kevin_row.correlation).alignment = wrap
        ws.cell(row=block_top, column=2, value=kevin_row.drawing).alignment = wrap
        ws.cell(row=block_top, column=3, value=kevin_row.revision).alignment = wrap
        ws.cell(row=block_top, column=4, value=kevin_row.detail).alignment = wrap

        scope_text = _format_scope_text(kevin_row.scope_lines)
        scope_cell = ws.cell(row=block_top, column=SCOPE_COL, value=scope_text)
        scope_cell.alignment = wrap
        if block_bottom > block_top:
            ws.merge_cells(
                start_row=block_top, start_column=SCOPE_COL,
                end_row=block_bottom, end_column=SCOPE_COL,
            )
            ws.merge_cells(
                start_row=block_top, start_column=DETAIL_VIEW_COL,
                end_row=block_bottom, end_column=DETAIL_VIEW_COL,
            )

        for r in range(block_top, block_bottom + 1):
            ws.row_dimensions[r].height = 18

        if kevin_row.crop_path:
            try:
                _embed_image(ws, kevin_row.crop_path, anchor_row=block_top, anchor_col=DETAIL_VIEW_COL)
            except Exception:
                ws.cell(row=block_top, column=DETAIL_VIEW_COL, value=f"[crop: {Path(kevin_row.crop_path).name}]")

        cursor = block_bottom + 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _format_scope_text(lines: list[str]) -> str:
    if not lines:
        return ""
    if len(lines) == 1:
        return lines[0]
    return "\n".join(f"{i}) {line}" for i, line in enumerate(lines, 1))


def _embed_image(ws, image_path: str, *, anchor_row: int, anchor_col: int) -> None:
    """Resize the source crop to fit the merged cell area and anchor it there.

    openpyxl's image insertion takes pixel dimensions; we resize through Pillow
    first so the embedded PNG is small (Kevin's file uses ~500px-wide crops).
    """
    target_w, target_h = CROP_TARGET_PX
    with PILImage.open(image_path) as src:
        src.load()
        src = src.convert("RGB") if src.mode in ("RGBA", "P") else src
        src.thumbnail((target_w, target_h), PILImage.LANCZOS)
        buf = io.BytesIO()
        src.save(buf, format="PNG", optimize=True)
        buf.seek(0)
    img = XLImage(buf)
    img.anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
    ws.add_image(img)
