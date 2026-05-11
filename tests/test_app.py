from __future__ import annotations

import io
import json
import os
import re
import threading
import time
from collections import Counter
from dataclasses import replace
from pathlib import Path

import fitz
import pytest

import backend.cli as cli_module
from backend.bulk_review_jobs import BulkReviewJobConflict, BulkReviewJobManager
from backend.cli import approve_cloudhammer_detections, main as cli_main
from backend.cloudhammer_client.inference import ManifestCloudInferenceClient
from backend.cloudhammer_client.live_pipeline import CloudHammerRunResult, LiveCloudHammerPipeline
from backend.cloudhammer_client.schemas import CloudDetection
from backend.crop_adjustments import CROP_ADJUSTMENT_KEY, crop_box_to_page_box, crop_adjustment_payload, selected_review_page_boxes
from backend.deliverables.excel_exporter import ExportBlockedError, Exporter
from backend.deliverables.review_packet import build_review_packet
from backend.geometry_corrections import GEOMETRY_CORRECTION_KEY, apply_geometry_correction
from backend.legend_context import (
    LEGEND_CONTEXT_KEY,
    classify_legend_context,
    enrich_workspace_legend_context,
    extract_symbol_definitions,
    legend_context_payload,
)
from backend.pre_review import (
    PRE_REVIEW_1,
    PRE_REVIEW_2,
    OpenAIPreReviewProvider,
    build_pre_review_provider_from_env,
    ensure_workspace_pre_review,
    normalize_pre_review_2,
    pre_review_payload,
    select_pre_review_source,
)
from backend.projects import ProjectRecord, ProjectRegistry
from backend.revision_state.models import ChangeItem, CloudCandidate, NarrativeEntry, RevisionSet, SheetVersion
from backend.revision_state.page_classification import sheet_is_index_like
from backend.review import change_item_needs_attention
from backend.review_events import record_bulk_review_updates, record_internal_review_event, record_review_update
from backend.review_queue import ensure_queue_order, ordered_change_items, visible_change_items
from backend.revision_state.tracker import RevisionScanner
from backend.scope_extraction import enrich_workspace_scope_text, extract_cloud_scope_text
from backend.utils import choose_best_sheet_id, parse_detail_ref
from backend.workspace import WorkspaceStore
from webapp.app import create_app, discipline_for_sheet


def write_minimal_drawing_pdf(path: Path, *, scope_text: str = "PROVIDE NEW GRAB BAR BLOCKING") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    document = fitz.open()
    page = document.new_page(width=600, height=800)
    page.insert_text((60, 80), f"AE101 FIRST FLOOR PLAN {scope_text}", fontsize=12)
    page.insert_text((430, 720), "AE101", fontsize=14)
    page.insert_text((430, 742), "FIRST FLOOR PLAN", fontsize=10)
    document.save(path)
    document.close()


def register_workspace_project(workspace_dir: Path, *, name: str = "Test Project") -> None:
    store = WorkspaceStore(workspace_dir).load()
    registry = ProjectRegistry(workspace_dir).load()
    registry.projects = [
        ProjectRecord(
            id="test-project",
            name=name,
            workspace_dir=str(workspace_dir.resolve()),
            input_dir=store.data.input_dir,
            status="active",
            created_at=store.data.created_at,
        )
    ]
    registry.save()


def csrf_token_from(response) -> str:
    match = re.search(rb'name="csrf_token" value="([^"]+)"', response.data)
    if not match:
        match = re.search(rb'<meta name="csrf-token" content="([^"]+)"', response.data)
    assert match
    return match.group(1).decode("utf-8")


def test_regression_fixture_metrics(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    expected = json.loads((Path(__file__).parent / "fixtures" / "expected_workspace_metrics.json").read_text(encoding="utf-8"))
    actual = {
        "revision_sets": len(store.data.revision_sets),
        "documents": len(store.data.documents),
        "sheets": len(store.data.sheets),
        "clouds": len(store.data.clouds),
        "change_items": len(store.data.change_items),
        "preflight_issues": len(store.data.preflight_issues),
        "attention_items": len([item for item in store.data.change_items if change_item_needs_attention(item)]),
        "visual_methods": dict(
            Counter(
                item.provenance.get("extraction_method", "narrative")
                for item in store.data.change_items
                if item.provenance.get("source") == "visual-region"
            )
        ),
    }
    assert actual == expected


def test_attention_helper_tolerates_malformed_extraction_signal():
    item = ChangeItem(
        id="change-1",
        sheet_version_id="sheet-1",
        cloud_candidate_id="cloud-1",
        sheet_id="AE101",
        detail_ref="Detail 1",
        raw_text="Scope",
        normalized_text="scope",
        provenance={"source": "visual-region", "extraction_signal": "not-a-number"},
    )

    assert not change_item_needs_attention(item)


def test_scan_generates_supersedence_and_ae113(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    assert len(store.data.revision_sets) == 2
    assert store.data.documents
    assert store.data.preflight_issues

    ae600_versions = [sheet for sheet in store.data.sheets if sheet.sheet_id == "AE600"]
    assert {sheet.status for sheet in ae600_versions} == {"active", "superseded"}

    assert store.data.clouds == []
    assert store.data.change_items
    assert {item.provenance["source"] for item in store.data.change_items} == {"narrative"}
    assert {item.sheet_id for item in store.data.change_items} >= {"AE107", "AE108", "AE109", "GI104"}


def test_export_does_not_block_without_attention_items(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    outputs = Exporter(store).export()
    assert "revision_changelog_xlsx" in outputs


def test_manifest_cloudhammer_client_returns_scaled_detections(tmp_path: Path):
    crop_path = tmp_path / "crop.png"
    crop_path.write_bytes(b"png")
    pdf_path = tmp_path / "revision.pdf"
    pdf_path.write_bytes(b"%PDF")
    manifest = tmp_path / "manifest.jsonl"
    manifest.write_text(
        json.dumps(
            {
                "candidate_id": "cand-1",
                "pdf_path": str(pdf_path),
                "page_number": 2,
                "page_width": 1000,
                "page_height": 500,
                "tight_crop_image_path": str(crop_path),
                "bbox_page_xywh": [100, 50, 200, 100],
                "whole_cloud_confidence": 0.91,
                "policy_bucket": "auto_deliverable_candidate",
            }
        )
        + "\n",
        encoding="utf-8",
    )
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf=str(pdf_path),
        page_number=2,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        render_path=str(tmp_path / "page.png"),
        width=500,
        height=250,
    )

    detections = ManifestCloudInferenceClient(manifest).detect(page=None, sheet=sheet)  # type: ignore[arg-type]

    assert len(detections) == 1
    assert detections[0].bbox == [50, 25, 100, 50]
    assert detections[0].image_path == str(crop_path.resolve())
    assert detections[0].extraction_method == "cloudhammer_manifest"


def test_manifest_cloudhammer_client_filters_release_and_bbox_noise(tmp_path: Path):
    pdf_path = tmp_path / "revision.pdf"
    pdf_path.write_bytes(b"%PDF")
    rows = [
        {
            "candidate_id": "valid",
            "pdf_path": str(pdf_path),
            "page_number": 1,
            "page_width": 100,
            "page_height": 50,
            "bbox_page_xyxy": [90, 40, 120, 80],
            "whole_cloud_confidence": 0.88,
            "policy_bucket": "auto_deliverable_candidate",
        },
        {
            "candidate_id": "rejected",
            "pdf_path": str(pdf_path),
            "page_number": 1,
            "page_width": 100,
            "page_height": 50,
            "bbox_page_xywh": [10, 10, 20, 20],
            "policy_bucket": "false_positive",
        },
        {
            "candidate_id": "invalid",
            "pdf_path": str(pdf_path),
            "page_number": 1,
            "page_width": 100,
            "page_height": 50,
            "bbox_page_xywh": [10, 10, -5, 20],
        },
        {
            "candidate_id": "missing-page",
            "pdf_path": str(pdf_path),
            "bbox_page_xywh": [10, 10, 20, 20],
        },
    ]
    manifest = tmp_path / "manifest.jsonl"
    manifest.write_text("\n".join(json.dumps(row) for row in rows) + "\n", encoding="utf-8")
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf=str(pdf_path),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        render_path=str(tmp_path / "page.png"),
        width=200,
        height=100,
    )

    client = ManifestCloudInferenceClient(manifest)
    detections = client.detect(page=None, sheet=sheet)  # type: ignore[arg-type]

    assert [item.metadata["cloudhammer_candidate_id"] for item in detections] == ["valid"]
    assert detections[0].bbox == [180, 80, 20, 20]
    assert client.stats["total_rows"] == 4
    assert client.stats["indexed_rows"] == 1
    assert client.stats["skipped_policy"] == 1
    assert client.stats["skipped_invalid_bbox"] == 1
    assert client.stats["skipped_missing_pdf_page"] == 1


def test_scan_reprocesses_cache_when_cloudhammer_manifest_changes(tmp_path: Path):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    pdf_path = input_dir / "Revision #1 - Test" / "drawing.pdf"
    write_minimal_drawing_pdf(pdf_path)

    initial = RevisionScanner(input_dir, workspace_dir).scan()
    assert initial.data.clouds == []

    crop_path = tmp_path / "crop.png"
    crop_path.write_bytes(b"png")
    manifest = tmp_path / "manifest.jsonl"
    manifest.write_text(
        json.dumps(
            {
                "candidate_id": "live-1",
                "pdf_path": str(pdf_path),
                "page_number": 1,
                "bbox_page_xywh": [70, 70, 120, 80],
                "whole_cloud_confidence": 0.93,
                "policy_bucket": "auto_deliverable_candidate",
                "crop_image_path": str(crop_path),
            }
        )
        + "\n",
        encoding="utf-8",
    )

    scanner = RevisionScanner(input_dir, workspace_dir, cloud_inference_client=ManifestCloudInferenceClient(manifest))
    refreshed = scanner.scan()

    assert scanner.cache_hits == 0
    assert len(refreshed.data.clouds) == 1
    assert refreshed.data.change_items[0].provenance["extraction_method"] == "cloudhammer_manifest"


def test_scan_records_invalid_pdf_as_diagnostic_without_crashing(tmp_path: Path):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    bad_pdf = input_dir / "Revision #1 - Bad Package" / "not-really.pdf"
    bad_pdf.parent.mkdir(parents=True)
    bad_pdf.write_bytes(b"not a valid pdf")

    store = RevisionScanner(input_dir, workspace_dir).scan()

    assert len(store.data.documents) == 1
    assert store.data.documents[0].page_count == 0
    assert store.data.documents[0].max_severity == "high"
    assert any(issue.code == "pdf_open_failed" for issue in store.data.preflight_issues)


def test_sheet_id_parser_prefers_repeated_plumbing_sheet_over_late_arch_reference():
    text = (
        "Drawing Number PL302 Existing 1st Floor Med Gas PL302 "
        "Refer to PL601 for symbols. Refer to drawing AE102 for phasing."
    )

    assert choose_best_sheet_id(text, preferred_prefixes=("PL", "P", "MP"), prefer_repeated=True) == "PL302"


def test_sheet_id_parser_supports_plain_plumbing_prefix():
    assert choose_best_sheet_id("Drawing Number P101 P101 Refer to AE102", preferred_prefixes=("PL", "P", "MP"), prefer_repeated=True) == "P101"
    assert parse_detail_ref("See 4/P101 for plumbing scope") == "4/P101"
    assert discipline_for_sheet("P101") == "Plumbing"


def test_plumbing_pdf_metadata_does_not_choose_late_arch_reference():
    pdf_path = Path.cwd() / "revision_sets" / "Revision #4 - Dental Air" / "260219 - VA Biloxi Rev 4_Plumbing 1.pdf"
    scanner = RevisionScanner.__new__(RevisionScanner)
    document = fitz.open(pdf_path)
    try:
        page = document[0]
        metadata = scanner._extract_sheet_metadata(page, page.get_text("text"), pdf_path)
    finally:
        document.close()

    assert metadata["sheet_id"] == "PL302"


def test_approve_cloudhammer_detections_only_marks_manifest_visual_items(tmp_path: Path):
    store = WorkspaceStore(tmp_path / "workspace").create(tmp_path)
    store.data.change_items = [
        ChangeItem(
            id="cloud",
            sheet_version_id="s1",
            cloud_candidate_id="c1",
            sheet_id="AE101",
            detail_ref=None,
            raw_text="CloudHammer detected revision cloud",
            normalized_text="cloudhammer detected revision cloud",
            provenance={"source": "visual-region", "extraction_method": "cloudhammer_manifest"},
        ),
        ChangeItem(
            id="narrative",
            sheet_version_id="s1",
            cloud_candidate_id=None,
            sheet_id="AE101",
            detail_ref=None,
            raw_text="Narrative item",
            normalized_text="narrative item",
            provenance={"source": "narrative"},
        ),
    ]

    changed = approve_cloudhammer_detections(store)

    assert changed == 1
    assert store.data.change_items[0].status == "approved"
    assert store.data.change_items[1].status == "pending"


def test_workspace_persists_portable_relative_paths(tmp_path: Path):
    workspace_dir = tmp_path / "workspace"
    input_dir = workspace_dir / "input"
    input_dir.mkdir(parents=True)
    store = WorkspaceStore(workspace_dir).create(input_dir)
    project_pdf = Path.cwd() / "revision_sets" / "Revision #1 - Drawing Changes" / "Revision #1 - Drawing Changes.pdf"
    render_path = store.page_path("sheet-1")
    render_path.write_bytes(b"png")
    crop_path = store.crop_path("crop-1")
    crop_path.write_bytes(b"png")
    store.data.revision_sets = [
        RevisionSet(
            id="rev-1",
            label="Revision #1 - Drawing Changes",
            source_dir=str(project_pdf.parent),
            set_number=1,
            set_date=None,
            pdf_paths=[str(project_pdf)],
        )
    ]
    store.data.sheets = [
        SheetVersion(
            id="sheet-1",
            revision_set_id="rev-1",
            source_pdf=str(project_pdf),
            page_number=1,
            sheet_id="AE101",
            sheet_title="Plan",
            issue_date=None,
            render_path=str(render_path),
        )
    ]
    store.data.clouds = [
        CloudCandidate(
            id="cloud-1",
            sheet_version_id="sheet-1",
            bbox=[0, 0, 10, 10],
            image_path=str(crop_path),
            page_image_path=str(render_path),
            confidence=0.9,
            extraction_method="test",
            nearby_text="",
            detail_ref=None,
        )
    ]

    store.save()

    raw = store.data_path.read_text(encoding="utf-8")
    assert str(tmp_path) not in raw
    assert str(Path.cwd()) not in raw
    assert "revision_sets/" in raw
    assert "assets/pages/sheet-1.png" in raw
    assert "input" in raw

    loaded = WorkspaceStore(workspace_dir).load()
    assert Path(loaded.data.input_dir).resolve() == input_dir.resolve()
    assert Path(loaded.data.sheets[0].source_pdf).resolve() == project_pdf.resolve()
    assert Path(loaded.data.sheets[0].render_path).resolve() == render_path.resolve()
    assert Path(loaded.data.clouds[0].image_path).resolve() == crop_path.resolve()


def test_visual_region_signal_score_handles_cloudhammer_placeholder():
    scanner = RevisionScanner.__new__(RevisionScanner)
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
    )

    score = scanner._signal_score("CloudHammer detected revision cloud. OCR/scope extraction is not wired yet.", sheet)

    assert score > 0.5


def test_scope_extraction_reads_pdf_text_near_cloud():
    document = fitz.open()
    page = document.new_page(width=300, height=200)
    page.insert_text((82, 95), "PROVIDE NEW GRAB BAR BLOCKING", fontsize=10)
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        width=300,
        height=200,
    )

    result = extract_cloud_scope_text(page, sheet, [70, 70, 95, 55])

    document.close()
    assert result.reason == "text-layer-near-cloud"
    assert result.signal > 0.7
    assert "PROVIDE NEW GRAB BAR BLOCKING" in result.text


def test_scope_extraction_keeps_nearby_text_local_to_cloud():
    document = fitz.open()
    page = document.new_page(width=300, height=220)
    page.insert_text((82, 95), "PROVIDE NEW GRAB BAR BLOCKING", fontsize=10)
    page.insert_text((10, 185), "REMOVE EXISTING FLOOR FINISH THROUGHOUT CORRIDOR AND PATCH WALLS", fontsize=10)
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        width=300,
        height=220,
    )

    result = extract_cloud_scope_text(page, sheet, [70, 70, 95, 55])

    document.close()
    assert "PROVIDE NEW GRAB BAR BLOCKING" in result.text
    assert "REMOVE EXISTING FLOOR FINISH" not in result.text
    assert result.context_bbox[1] > 40
    assert result.context_bbox[3] < 130


def test_scope_extraction_filters_isolated_numeric_noise_but_keeps_tags():
    document = fitz.open()
    page = document.new_page(width=300, height=220)
    page.insert_text((82, 95), "7 Z.8 PROVIDE NEW EXHAUST FAN", fontsize=10)
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        width=300,
        height=220,
    )

    result = extract_cloud_scope_text(page, sheet, [70, 70, 95, 55])

    document.close()
    assert "PROVIDE NEW EXHAUST FAN" in result.text
    assert "Z.8" in result.text
    assert " 7 " not in f" {result.text} "


def test_scope_extraction_flags_cloud_without_readable_text():
    document = fitz.open()
    page = document.new_page(width=300, height=200)
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        width=300,
        height=200,
    )

    result = extract_cloud_scope_text(page, sheet, [70, 70, 95, 55])

    document.close()
    assert result.reason == "no-readable-text"
    assert result.signal < 0.3
    assert "No readable scope text" in result.text


def build_pre_review_test_store(tmp_path: Path) -> WorkspaceStore:
    from PIL import Image

    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    pdf_path = input_dir / "Revision #1" / "drawing.pdf"
    pdf_path.parent.mkdir(parents=True)
    write_minimal_drawing_pdf(pdf_path, scope_text="PROVIDE NEW ROOF CURB")
    store = WorkspaceStore(workspace_dir).create(input_dir)
    crop_path = store.crop_path("cloud-1")
    render_path = store.page_path("sheet-1")
    Image.new("RGB", (200, 120), "white").save(crop_path)
    Image.new("RGB", (400, 240), "white").save(render_path)
    store.data.revision_sets = [
        RevisionSet(
            id="rev-1",
            label="Revision #1",
            source_dir=str(pdf_path.parent),
            set_number=1,
            set_date=None,
            pdf_paths=[str(pdf_path)],
        )
    ]
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf=str(pdf_path),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        render_path=str(render_path),
        width=400,
        height=240,
        status="active",
    )
    cloud = CloudCandidate(
        id="cloud-1",
        sheet_version_id=sheet.id,
        bbox=[80, 40, 120, 80],
        image_path=str(crop_path),
        page_image_path="",
        confidence=0.91,
        extraction_method="cloudhammer_manifest",
        nearby_text="Cloud Only - PROVIDE NEW ROOF CURB",
        detail_ref=None,
        scope_text="Cloud Only - PROVIDE NEW ROOF CURB",
        scope_reason="text-layer-near-cloud",
        scope_signal=0.78,
        scope_method="pdf-text-layer",
        metadata={
            "bbox_page_xywh": [80, 40, 120, 80],
            "bbox_page_xyxy": [80, 40, 200, 120],
            "crop_box_page_xywh": [40, 20, 240, 160],
            "crop_box_page_xyxy": [40, 20, 280, 180],
            "page_width": 400,
            "page_height": 240,
        },
    )
    item = ChangeItem(
        id="change-1",
        sheet_version_id=sheet.id,
        cloud_candidate_id=cloud.id,
        sheet_id=sheet.sheet_id,
        detail_ref=None,
        raw_text=cloud.scope_text,
        normalized_text=cloud.scope_text.lower(),
        provenance={"source": "visual-region", "extraction_method": "cloudhammer_manifest"},
    )
    store.data.sheets = [sheet]
    store.data.clouds = [cloud]
    store.data.change_items = [item]
    store.save()
    return store


def append_pre_review_test_item(store: WorkspaceStore, index: int) -> None:
    from PIL import Image

    base_cloud = store.data.clouds[0]
    base_item = store.data.change_items[0]
    crop_path = store.crop_path(f"cloud-{index}")
    Image.new("RGB", (200, 120), "white").save(crop_path)
    cloud = replace(base_cloud, id=f"cloud-{index}", image_path=str(crop_path))
    item = replace(base_item, id=f"change-{index}", cloud_candidate_id=cloud.id)
    store.data.clouds.append(cloud)
    store.data.change_items.append(item)


def test_old_change_items_load_with_queue_defaults(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    payload = json.loads(store.data_path.read_text(encoding="utf-8"))
    for item in payload["change_items"]:
        item.pop("queue_order", None)
        item.pop("parent_change_item_id", None)
        item.pop("superseded_by_change_item_ids", None)
        item.pop("superseded_reason", None)
        item.pop("superseded_at", None)
    store.data_path.write_text(json.dumps(payload), encoding="utf-8")

    loaded = WorkspaceStore(store.workspace_dir).load()

    assert loaded.data.change_items[0].queue_order == 0.0
    assert loaded.data.change_items[0].parent_change_item_id is None
    assert loaded.data.change_items[0].superseded_by_change_item_ids == []
    ordered, changed = ensure_queue_order(loaded.data.change_items)
    assert changed is True
    assert ordered[0].queue_order > 0


def test_review_queue_order_and_superseded_items_are_hidden(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    append_pre_review_test_item(store, 3)
    store.data.change_items = [
        replace(store.data.change_items[0], queue_order=3000),
        replace(
            store.data.change_items[1],
            queue_order=1000,
            superseded_by_change_item_ids=["change-3"],
            superseded_reason="overmerge_split",
            superseded_at="2026-05-11T00:00:00+00:00",
        ),
        replace(store.data.change_items[2], queue_order=2000),
    ]
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir, bulk_review_manager=BulkReviewJobManager(run_async=False))
    client = app.test_client()

    queue = client.get("/changes?status=all")
    detail = client.get("/changes/change-3?queue=all")

    assert queue.status_code == 200
    assert b"change-2" not in queue.data
    assert queue.data.find(b"change-3") < queue.data.find(b"change-1")
    assert detail.status_code == 200
    assert b"1 / 2" in detail.data


def test_bulk_review_ignores_superseded_parents(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.data.change_items = [
        replace(
            store.data.change_items[0],
            queue_order=1000,
            superseded_by_change_item_ids=["change-2"],
            superseded_reason="partial_correction",
            superseded_at="2026-05-11T00:00:00+00:00",
        ),
        replace(store.data.change_items[1], queue_order=1000.001),
    ]
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir, bulk_review_manager=BulkReviewJobManager(run_async=False))
    client = app.test_client()

    response = client.post(
        "/changes/bulk-review",
        data={"change_ids": ["change-1", "change-2"], "status": "approved", "redirect_to": "/changes"},
    )
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert response.status_code == 302
    parent = loaded.get_change_item("change-1")
    child = loaded.get_change_item("change-2")
    assert parent.status == "pending"
    assert child.status == "approved"
    assert [event.change_item_id for event in loaded.data.review_events] == ["change-2"]


def test_record_bulk_review_updates_saves_once_and_creates_review_events(tmp_path: Path, monkeypatch):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    append_pre_review_test_item(store, 3)
    store.save()
    save_count = 0
    original_save = store.save

    def counted_save():
        nonlocal save_count
        save_count += 1
        original_save()

    monkeypatch.setattr(store, "save", counted_save)

    result = record_bulk_review_updates(
        store,
        project_id="test-project",
        item_changes={
            "change-1": {"status": "approved", "reviewer_text": "Scope 1"},
            "change-2": {"status": "approved", "reviewer_text": "Scope 2"},
            "change-3": {"status": "approved", "reviewer_text": "Scope 3"},
        },
        reviewer_id="reviewer@example.com",
        review_session_id="session-1",
        action="accept",
    )

    loaded = WorkspaceStore(store.workspace_dir).load()
    assert result.updated_count == 3
    assert save_count == 1
    assert [item.status for item in loaded.data.change_items] == ["approved", "approved", "approved"]
    assert [event.change_item_id for event in loaded.data.review_events] == ["change-1", "change-2", "change-3"]
    assert all(event.action == "accept" for event in loaded.data.review_events)


def test_bulk_review_route_accept_all_writes_workspace_once(tmp_path: Path, monkeypatch):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    append_pre_review_test_item(store, 3)
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir, bulk_review_manager=BulkReviewJobManager(run_async=False))
    client = app.test_client()
    save_count = 0
    original_save = WorkspaceStore.save

    def counted_save(self):
        nonlocal save_count
        save_count += 1
        return original_save(self)

    monkeypatch.setattr(WorkspaceStore, "save", counted_save)

    response = client.post(
        "/changes/bulk-review",
        data={
            "change_ids": ["change-1", "change-2", "change-3"],
            "status": "approved",
            "redirect_to": "/changes",
        },
    )
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert response.status_code == 302
    assert save_count == 1
    assert [item.status for item in loaded.data.change_items] == ["approved", "approved", "approved"]
    assert [event.change_item_id for event in loaded.data.review_events] == ["change-1", "change-2", "change-3"]


def test_bulk_review_job_manager_skips_ineligible_items_and_exports_status(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    append_pre_review_test_item(store, 3)
    raw_text = store.data.change_items[2].raw_text
    store.data.change_items = [
        store.data.change_items[0],
        replace(
            store.data.change_items[1],
            superseded_by_change_item_ids=["change-1"],
            superseded_reason="partial_correction",
            superseded_at="2026-05-11T00:00:00+00:00",
        ),
        replace(store.data.change_items[2], status="approved", reviewer_text=raw_text),
    ]
    store.save()
    manager = BulkReviewJobManager(run_async=False)

    job = manager.start_job(
        project_id="test-project",
        workspace_dir=store.workspace_dir,
        selected_change_ids=["change-1", "change-1", "change-2", "missing", "change-3"],
        requested_status="approved",
        reviewer_id="reviewer@example.com",
        review_session_id="session-1",
    )
    payload = manager.status_payload("test-project", job.id)
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert job.state == "done"
    assert payload["total_selected"] == 5
    assert payload["eligible_count"] == 1
    assert payload["updated_count"] == 1
    assert payload["skipped_count"] == 4
    assert loaded.get_change_item("change-1").status == "approved"
    assert loaded.get_change_item("change-2").status == "pending"
    assert [event.change_item_id for event in loaded.data.review_events] == ["change-1"]


def test_bulk_review_job_manager_rejects_second_active_job(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    started = threading.Event()
    release = threading.Event()

    def blocking_store_factory(path: Path) -> WorkspaceStore:
        started.set()
        release.wait(timeout=5)
        return WorkspaceStore(path)

    manager = BulkReviewJobManager(store_factory=blocking_store_factory)
    job = manager.start_job(
        project_id="test-project",
        workspace_dir=store.workspace_dir,
        selected_change_ids=["change-1"],
        requested_status="approved",
        reviewer_id=None,
        review_session_id=None,
    )
    assert started.wait(timeout=5)

    with pytest.raises(BulkReviewJobConflict):
        manager.start_job(
            project_id="test-project",
            workspace_dir=store.workspace_dir,
            selected_change_ids=["change-1"],
            requested_status="rejected",
            reviewer_id=None,
            review_session_id=None,
        )

    release.set()
    assert manager.wait(job.id, timeout=5)


def test_bulk_review_job_manager_records_failure_state(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)

    def failing_store_factory(path: Path) -> WorkspaceStore:
        raise RuntimeError("workspace unavailable")

    manager = BulkReviewJobManager(store_factory=failing_store_factory, run_async=False)

    job = manager.start_job(
        project_id="test-project",
        workspace_dir=store.workspace_dir,
        selected_change_ids=["change-1"],
        requested_status="approved",
        reviewer_id=None,
        review_session_id=None,
    )

    assert job.state == "failed"
    assert "workspace unavailable" in job.error
    assert manager.status_payload("test-project", job.id)["state"] == "failed"


def test_bulk_review_route_json_starts_job_and_status_endpoint_reports_completion(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.save()
    register_workspace_project(store.workspace_dir)
    manager = BulkReviewJobManager(run_async=False)
    app = create_app(store.workspace_dir, bulk_review_manager=manager)
    client = app.test_client()

    response = client.post(
        "/changes/bulk-review",
        data={"change_ids": ["change-1", "change-2"], "status": "approved", "redirect_to": "/changes"},
        headers={"Accept": "application/json"},
    )
    payload = response.get_json()
    status = client.get(payload["status_url"])
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert response.status_code == 202
    assert payload["job_id"]
    assert status.status_code == 200
    assert status.get_json()["state"] == "done"
    assert [item.status for item in loaded.data.change_items] == ["approved", "approved"]
    assert [event.action for event in loaded.data.review_events] == ["accept", "accept"]


def test_bulk_review_running_allows_get_navigation_and_blocks_mutations(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    started = threading.Event()
    release = threading.Event()

    def blocking_store_factory(path: Path) -> WorkspaceStore:
        started.set()
        release.wait(timeout=5)
        return WorkspaceStore(path)

    manager = BulkReviewJobManager(store_factory=blocking_store_factory)
    job = manager.start_job(
        project_id="test-project",
        workspace_dir=store.workspace_dir,
        selected_change_ids=["change-1"],
        requested_status="approved",
        reviewer_id=None,
        review_session_id=None,
    )
    assert started.wait(timeout=5)
    app = create_app(store.workspace_dir, bulk_review_manager=manager)
    client = app.test_client()

    get_paths = ["/changes", "/changes/change-1", "/overview", "/sheets/sheet-1", "/export"]
    get_responses = [client.get(path) for path in get_paths]
    review_response = client.post(
        "/changes/change-1/review",
        data={"status": "approved", "reviewer_text": "Approved"},
        headers={"Accept": "application/json"},
    )
    export_response = client.post("/export/run", headers={"Accept": "application/json"})
    duplicate_bulk = client.post(
        "/changes/bulk-review",
        data={"change_ids": ["change-1"], "status": "rejected", "redirect_to": "/changes"},
        headers={"Accept": "application/json"},
    )

    release.set()
    assert manager.wait(job.id, timeout=5)
    assert [response.status_code for response in get_responses] == [200, 200, 200, 200, 200]
    assert b"data-bulk-review-status" in get_responses[0].data
    assert review_response.status_code == 409
    assert export_response.status_code == 409
    assert duplicate_bulk.status_code == 409


def test_workspace_save_writes_valid_json_atomically(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    store.data.change_items = [replace(store.data.change_items[0], status="approved")]

    store.save()

    payload = json.loads(store.data_path.read_text(encoding="utf-8"))
    assert payload["change_items"][0]["status"] == "approved"
    assert list(store.workspace_dir.glob(".workspace.json.*.tmp")) == []


def test_superseded_parents_are_hidden_from_exports_and_packet(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.data.change_items = [
        replace(
            store.data.change_items[0],
            status="approved",
            reviewer_text="Hidden parent",
            queue_order=1000,
            superseded_by_change_item_ids=["change-2"],
            superseded_reason="overmerge_split",
            superseded_at="2026-05-11T00:00:00+00:00",
        ),
        replace(store.data.change_items[1], status="approved", reviewer_text="Visible child", queue_order=1000.001),
    ]
    store.save()

    outputs = Exporter(store).export(force_attention=True)
    approved_rows = json.loads(Path(outputs["approved_changes_json"]).read_text(encoding="utf-8"))
    candidates = json.loads((store.output_dir / "pricing_change_candidates.json").read_text(encoding="utf-8"))
    packet = build_review_packet(store)

    assert [row["change_id"] for row in approved_rows] == ["change-2"]
    assert all(row["change_id"] != "change-1" for row in candidates)
    assert packet.item_count == 1


def test_queue_supersession_metadata_survives_rescan(tmp_path: Path):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    pdf_path = input_dir / "Revision #1 - Test" / "drawing.pdf"
    write_minimal_drawing_pdf(pdf_path)
    crop_path = tmp_path / "crop.png"
    crop_path.write_bytes(b"png")
    manifest = tmp_path / "manifest.jsonl"
    manifest.write_text(
        json.dumps(
            {
                "candidate_id": "live-1",
                "pdf_path": str(pdf_path),
                "page_number": 1,
                "bbox_page_xywh": [70, 70, 120, 80],
                "whole_cloud_confidence": 0.93,
                "policy_bucket": "auto_deliverable_candidate",
                "crop_image_path": str(crop_path),
            }
        )
        + "\n",
        encoding="utf-8",
    )
    initial = RevisionScanner(input_dir, workspace_dir, cloud_inference_client=ManifestCloudInferenceClient(manifest)).scan()
    parent = replace(
        initial.data.change_items[0],
        queue_order=1000,
        superseded_by_change_item_ids=["change-child"],
        superseded_reason="partial_correction",
        superseded_at="2026-05-11T00:00:00+00:00",
    )
    child_cloud = replace(initial.data.clouds[0], id="cloud-child", bbox=[75, 75, 100, 60])
    child = replace(
        parent,
        id="change-child",
        cloud_candidate_id=child_cloud.id,
        parent_change_item_id=parent.id,
        superseded_by_change_item_ids=[],
        superseded_reason=None,
        superseded_at=None,
        queue_order=1000.001,
        provenance={**parent.provenance, "scopeledger.geometry_correction.v1": {"mode": "partial"}},
    )
    initial.data.clouds.append(child_cloud)
    initial.data.change_items = [parent, child]
    initial.save()

    rescanned = RevisionScanner(input_dir, workspace_dir, cloud_inference_client=ManifestCloudInferenceClient(manifest)).scan()

    restored_parent = next(item for item in rescanned.data.change_items if item.id == parent.id)
    restored_child = next(item for item in rescanned.data.change_items if item.id == child.id)
    assert restored_parent.superseded_by_change_item_ids == ["change-child"]
    assert restored_parent.queue_order == 1000
    assert restored_child.parent_change_item_id == parent.id
    assert any(cloud.id == "cloud-child" for cloud in rescanned.data.clouds)


def test_overmerge_split_creates_child_items_and_review_event(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    store.data.change_items = [replace(store.data.change_items[0], queue_order=1000)]
    store.save()
    parent = store.data.change_items[0]
    parent_cloud = store.data.clouds[0]
    sheet = store.data.sheets[0]

    result = apply_geometry_correction(
        store,
        parent,
        parent_cloud,
        sheet,
        mode="overmerge",
        crop_boxes=[[10, 10, 40, 30], [80, 40, 50, 35]],
        project_id="test-project",
        reviewer_id="reviewer@example.com",
        review_session_id="session-1",
    )
    loaded = WorkspaceStore(store.workspace_dir).load()

    loaded_parent = loaded.get_change_item(parent.id)
    assert loaded_parent.superseded_reason == "overmerge_split"
    assert loaded_parent.superseded_by_change_item_ids == [item.id for item in result.child_items]
    assert [item.id for item in visible_change_items(loaded.data.change_items)] == [item.id for item in result.child_items]
    assert all(item.parent_change_item_id == parent.id and item.status == "pending" for item in result.child_items)
    assert all(Path(cloud.image_path).exists() for cloud in result.child_clouds)
    assert loaded.get_cloud(parent_cloud.id).bbox == [80, 40, 120, 80]
    assert len(loaded.data.review_events) == 1
    event = loaded.data.review_events[0]
    assert event.action == "split"
    assert event.original_candidate_json["cloud_candidate"]["bbox"] == [80, 40, 120, 80]
    assert len(event.human_result_json["split_child_geometries"]) == 2


def test_partial_correction_route_creates_one_replacement_and_redirects(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    store.data.change_items = [replace(store.data.change_items[0], queue_order=1000)]
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.post(
        "/changes/change-1/geometry-correction",
        json={
            "mode": "partial",
            "crop_boxes": [[30, 20, 70, 50]],
            "queue_status": "pending",
            "search_query": "",
            "attention_only": "0",
        },
    )
    payload = response.get_json()
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert response.status_code == 200
    assert payload["redirect_url"].startswith("/changes/")
    parent = loaded.get_change_item("change-1")
    assert parent.superseded_reason == "partial_correction"
    assert len(parent.superseded_by_change_item_ids) == 1
    child = loaded.get_change_item(parent.superseded_by_change_item_ids[0])
    child_cloud = loaded.get_cloud(child.cloud_candidate_id)
    assert child.parent_change_item_id == "change-1"
    assert child.queue_order > parent.queue_order
    assert GEOMETRY_CORRECTION_KEY in child.provenance
    assert selected_review_page_boxes(child, child_cloud) == child.provenance[GEOMETRY_CORRECTION_KEY]["page_boxes"]
    assert loaded.data.review_events[0].action == "resize"


def test_geometry_correction_route_rejects_invalid_payload_without_mutation(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.post(
        "/changes/change-1/geometry-correction",
        json={"mode": "overmerge", "crop_boxes": [[30, 20, 70, 50]]},
    )
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert response.status_code == 400
    assert loaded.get_change_item("change-1").superseded_by_change_item_ids == []
    assert loaded.data.review_events == []
    assert len(loaded.data.change_items) == 1


def test_geometry_correction_children_export_as_normal_review_items(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    parent = replace(store.data.change_items[0], queue_order=1000)
    store.data.change_items = [parent]
    store.save()
    result = apply_geometry_correction(
        store,
        parent,
        store.data.clouds[0],
        store.data.sheets[0],
        mode="partial",
        crop_boxes=[[30, 20, 70, 50]],
        project_id="test-project",
        reviewer_id=None,
        review_session_id=None,
    )
    child = replace(result.child_items[0], status="approved", reviewer_text="Corrected child scope")
    store.update_change_item(child.id, status=child.status, reviewer_text=child.reviewer_text)

    outputs = Exporter(store).export(force_attention=True)
    approved_rows = json.loads(Path(outputs["approved_changes_json"]).read_text(encoding="utf-8"))
    packet = build_review_packet(store)

    assert [row["change_id"] for row in approved_rows] == [child.id]
    assert approved_rows[0]["text"] == "Corrected child scope"
    assert packet.item_count == 1


def test_superseded_parent_detail_redirects_and_review_actions_are_blocked(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    result = apply_geometry_correction(
        store,
        store.data.change_items[0],
        store.data.clouds[0],
        store.data.sheets[0],
        mode="partial",
        crop_boxes=[[30, 20, 70, 50]],
        project_id="test-project",
        reviewer_id=None,
        review_session_id=None,
    )
    child_id = result.child_items[0].id
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    detail = client.get("/changes/change-1?queue=pending")
    review = client.post(
        "/changes/change-1/review",
        data={
            "status": "approved",
            "queue_status": "pending",
            "search_query": "",
            "attention_only": "0",
        },
    )
    crop_adjustment = client.post("/changes/change-1/crop-adjustment", json={"crop_box": [20, 20, 40, 40]})
    geometry = client.post("/changes/change-1/geometry-correction", json={"mode": "partial", "crop_boxes": [[20, 20, 40, 40]]})
    loaded = WorkspaceStore(store.workspace_dir).load()

    assert detail.status_code == 302
    assert f"/changes/{child_id}".encode("utf-8") in detail.headers["Location"].encode("utf-8")
    assert review.status_code == 302
    assert f"/changes/{child_id}".encode("utf-8") in review.headers["Location"].encode("utf-8")
    assert crop_adjustment.status_code == 400
    assert geometry.status_code == 400
    assert loaded.get_change_item("change-1").status == "pending"
    assert len(loaded.data.review_events) == 1


def test_cli_approve_cloudhammer_detections_skips_superseded_items(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.data.change_items = [
        replace(
            store.data.change_items[0],
            superseded_by_change_item_ids=["change-2"],
            superseded_reason="partial_correction",
            superseded_at="2026-05-11T00:00:00+00:00",
        ),
        store.data.change_items[1],
    ]
    store.save()

    changed = approve_cloudhammer_detections(store)

    assert changed == 1
    assert store.get_change_item("change-1").status == "pending"
    assert store.get_change_item("change-2").status == "approved"


def test_scope_enrichment_skips_superseded_parent_items(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    parent = replace(
        store.data.change_items[0],
        raw_text="Keep parent text",
        normalized_text="keep parent text",
        reviewer_text="Keep parent text",
        superseded_by_change_item_ids=["change-child"],
        superseded_reason="partial_correction",
        superseded_at="2026-05-11T00:00:00+00:00",
    )
    store.data.change_items = [parent]
    store.save()

    enrich_workspace_scope_text(store, force=True)

    assert store.get_change_item("change-1").raw_text == "Keep parent text"
    assert store.get_change_item("change-1").reviewer_text == "Keep parent text"


def test_geometry_correction_controls_do_not_expose_internal_language(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.get("/changes/change-1")
    body = response.data.decode("utf-8").lower()

    assert response.status_code == 200
    assert "correct overmerge" in body
    assert "correct partial" in body
    for forbidden in ("training", "labeling", "mining", "eval", "cloudhammer", "gpt"):
        assert forbidden not in body


class FakePreReviewProvider:
    name = "fake_pre_review"
    enabled = True
    disabled_reason = ""

    def __init__(self):
        self.calls = 0

    def review(self, context):
        self.calls += 1
        return normalize_pre_review_2(
            {
                "geometry_decision": "adjusted_box",
                "boxes": [[45, 25, 110, 55]],
                "refined_text": "Provide new roof curb.",
                "reason": "Tighter around the clouded roof curb note.",
                "confidence": 0.82,
                "tags": ["tightened"],
            },
            context,
        )


def test_pre_review_success_stores_second_pass_without_hiding_candidate(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    provider = FakePreReviewProvider()

    summary = ensure_workspace_pre_review(store, provider)

    assert summary.total_count == 1
    assert summary.pre_review_2_count == 1
    assert len(store.data.change_items) == 1
    payload = pre_review_payload(store.data.change_items[0])
    assert payload["selected"] == PRE_REVIEW_1
    assert payload[PRE_REVIEW_2]["available"] is True
    assert payload[PRE_REVIEW_2]["text"] == "Provide new roof curb."
    assert provider.calls == 1


def test_pre_review_failure_continues_with_first_pass(tmp_path: Path):
    class FailingProvider:
        name = "failing_pre_review"
        enabled = True
        disabled_reason = ""

        def review(self, context):
            raise RuntimeError("temporary API failure")

    store = build_pre_review_test_store(tmp_path)

    summary = ensure_workspace_pre_review(store, FailingProvider())

    payload = pre_review_payload(store.data.change_items[0])
    assert summary.failed_count == 1
    assert summary.request_count == 1
    assert payload["selected"] == PRE_REVIEW_1
    assert payload[PRE_REVIEW_2]["available"] is False


def test_pre_review_selection_sets_reviewer_truth(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    ensure_workspace_pre_review(store, FakePreReviewProvider())
    item = store.data.change_items[0]

    selected = select_pre_review_source(item, PRE_REVIEW_2)

    payload = pre_review_payload(selected)
    assert payload["selected"] == PRE_REVIEW_2
    assert selected.reviewer_text == "Provide new roof curb."


def test_openai_pre_review_provider_uses_cache(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    store = build_pre_review_test_store(tmp_path)
    provider = OpenAIPreReviewProvider(model="gpt-5.5")
    calls = {"count": 0}

    def fake_call(context, input_image):
        calls["count"] += 1
        return json.dumps(
            {
                "geometry_decision": "same_box",
                "boxes": [[33, 15, 100, 60]],
                "refined_text": "Cached text.",
                "reason": "Same visible region.",
                "confidence": 0.7,
                "tags": ["same"],
            }
        )

    monkeypatch.setattr(provider, "_call_openai", fake_call)

    first = ensure_workspace_pre_review(store, provider, force=True)
    second = ensure_workspace_pre_review(store, provider, force=True)

    assert first.pre_review_2_count == 1
    assert second.cache_hits == 1
    assert calls["count"] == 1


def test_openai_pre_review_provider_batches_and_logs_usage(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    store = build_pre_review_test_store(tmp_path)
    for index in range(2, 4):
        append_pre_review_test_item(store, index)
    store.save()
    provider = OpenAIPreReviewProvider(model="gpt-5.5", batch_size=5)
    calls = {"count": 0}

    def fake_batch_call(contexts, input_images):
        calls["count"] += 1
        assert len(contexts) == 3
        assert len(input_images) == 3
        return {
            "text": json.dumps(
                {
                    "results": [
                        {
                            "item_id": context.item.id,
                            "geometry_decision": "same_box",
                            "boxes": [[33, 15, 100, 60]],
                            "refined_text": f"Batch text {context.item.id}.",
                            "reason": "Same visible region.",
                            "confidence": 0.7,
                            "tags": ["same"],
                        }
                        for context in contexts
                    ]
                }
            ),
            "usage": {
                "input_tokens": 90,
                "output_tokens": 30,
                "total_tokens": 120,
                "input_tokens_details": {"cached_tokens": 15},
            },
            "meta": {"duration_seconds": 1.2, "retry_count": 0},
        }

    monkeypatch.setattr(provider, "_call_openai_batch", fake_batch_call)

    summary = ensure_workspace_pre_review(store, provider, force=True)

    assert calls["count"] == 1
    assert summary.pre_review_2_count == 3
    assert summary.request_count == 1
    assert summary.input_tokens == 90
    assert summary.output_tokens == 30
    assert summary.total_tokens == 120
    assert summary.cached_input_tokens == 15
    cache_files = list((store.output_dir / "pre_review" / "cache").glob("*.json"))
    assert len(cache_files) == 3
    first_cache = json.loads(cache_files[0].read_text(encoding="utf-8"))
    assert first_cache["prompt_version"] == "scopeledger_pre_review_batch_prompt_v1"
    assert first_cache["usage"]["input_tokens"] > 0
    usage_lines = (store.output_dir / "pre_review" / "usage" / "pre_review_usage.jsonl").read_text(encoding="utf-8").splitlines()
    assert len(usage_lines) == 1
    assert json.loads(usage_lines[0])["batch_size"] == 3

    second = ensure_workspace_pre_review(store, provider, force=True)
    assert calls["count"] == 1
    assert second.cache_hits == 3
    assert second.request_count == 0
    assert second.total_tokens == 0


def test_batched_pre_review_reuses_existing_single_item_cache(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    store = build_pre_review_test_store(tmp_path)
    provider = OpenAIPreReviewProvider(model="gpt-5.5", batch_size=5)
    calls = {"single": 0, "batch": 0}

    def fake_single_call(context, input_image):
        calls["single"] += 1
        return json.dumps(
            {
                "geometry_decision": "same_box",
                "boxes": [[33, 15, 100, 60]],
                "refined_text": "Single cache text.",
                "reason": "Same visible region.",
                "confidence": 0.7,
                "tags": ["same"],
            }
        )

    monkeypatch.setattr(provider, "_call_openai", fake_single_call)
    first = ensure_workspace_pre_review(store, provider, force=True)
    assert first.pre_review_2_count == 1

    append_pre_review_test_item(store, 2)
    store.save()

    def fake_batch_call(contexts, input_images):
        calls["batch"] += 1
        assert [context.item.id for context in contexts] == ["change-2"]
        return {
            "text": json.dumps(
                {
                    "results": [
                        {
                            "item_id": "change-2",
                            "geometry_decision": "same_box",
                            "boxes": [[33, 15, 100, 60]],
                            "refined_text": "Batch text.",
                            "reason": "Same visible region.",
                            "confidence": 0.7,
                            "tags": ["same"],
                        }
                    ]
                }
            ),
            "usage": {"input_tokens": 20, "output_tokens": 10, "total_tokens": 30},
            "meta": {},
        }

    monkeypatch.setattr(provider, "_call_openai_batch", fake_batch_call)
    second = ensure_workspace_pre_review(store, provider, force=True)

    assert calls == {"single": 1, "batch": 1}
    assert second.pre_review_2_count == 2
    assert second.cache_hits == 1
    assert second.request_count == 1
    assert second.total_tokens == 30


def test_batched_pre_review_missing_or_duplicate_rows_fail_safely(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.save()
    provider = OpenAIPreReviewProvider(model="gpt-5.5", batch_size=5)

    def fake_batch_call(contexts, input_images):
        row = {
            "item_id": "change-1",
            "geometry_decision": "same_box",
            "boxes": [[33, 15, 100, 60]],
            "refined_text": "Duplicate should fail.",
            "reason": "Duplicate row.",
            "confidence": 0.7,
            "tags": ["duplicate"],
        }
        return {"text": json.dumps({"results": [row, row]}), "usage": {"total_tokens": 20}, "meta": {}}

    monkeypatch.setattr(provider, "_call_openai_batch", fake_batch_call)

    summary = ensure_workspace_pre_review(store, provider, force=True)

    assert summary.failed_count == 2
    assert summary.pre_review_2_count == 0
    loaded = WorkspaceStore(store.workspace_dir).load()
    assert all(pre_review_payload(item)["status"] == "failed" for item in loaded.data.change_items)


def test_batched_pre_review_progress_callback_saves_after_each_batch(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    for index in range(2, 5):
        append_pre_review_test_item(store, index)
    store.save()

    class BatchPreReviewProvider:
        name = "batch_fake"
        enabled = True
        disabled_reason = ""
        batch_size = 2

        def __init__(self):
            self.batch_lengths: list[int] = []

        def review(self, context):
            raise AssertionError("single-item review should not be used")

        def review_batch(self, contexts):
            self.batch_lengths.append(len(contexts))
            return {
                context.item.id: normalize_pre_review_2(
                    {
                        "geometry_decision": "same_box",
                        "boxes": [[33, 15, 100, 60]],
                        "refined_text": f"Progress text {context.item.id}.",
                        "reason": "Same visible region.",
                        "confidence": 0.7,
                        "tags": ["same"],
                    },
                    context,
                )
                for context in contexts
            }

    provider = BatchPreReviewProvider()
    snapshots: list[dict[str, object]] = []

    summary = ensure_workspace_pre_review(store, provider, force=True, progress_callback=lambda value: snapshots.append(value.to_status()))

    assert provider.batch_lengths == [2, 2]
    assert summary.pre_review_2_count == 4
    assert summary.request_count == 2
    assert snapshots
    assert snapshots[-1]["pre_review_2_count"] == 4
    loaded = WorkspaceStore(store.workspace_dir).load()
    assert all(pre_review_payload(item)[PRE_REVIEW_2]["available"] for item in loaded.data.change_items)


def test_pre_review_batch_size_env_defaults_accepts_and_clamps(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SCOPELEDGER_PREREVIEW_ENABLED", "1")
    monkeypatch.setenv("OPENAI_API_KEY", "test-key")
    monkeypatch.delenv("SCOPELEDGER_PREREVIEW_BATCH_SIZE", raising=False)
    default_provider = build_pre_review_provider_from_env()
    assert getattr(default_provider, "batch_size") == 5

    monkeypatch.setenv("SCOPELEDGER_PREREVIEW_BATCH_SIZE", "1")
    one_provider = build_pre_review_provider_from_env()
    assert getattr(one_provider, "batch_size") == 1

    monkeypatch.setenv("SCOPELEDGER_PREREVIEW_BATCH_SIZE", "100")
    clamped_provider = build_pre_review_provider_from_env()
    assert getattr(clamped_provider, "batch_size") == 10

    monkeypatch.setenv("SCOPELEDGER_PREREVIEW_BATCH_SIZE", "bad")
    with pytest.raises(RuntimeError, match="SCOPELEDGER_PREREVIEW_BATCH_SIZE"):
        build_pre_review_provider_from_env()


def test_review_route_persists_pre_review_selection_without_internal_labels(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    ensure_workspace_pre_review(store, FakePreReviewProvider())
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    page = client.get("/changes/change-1")

    assert page.status_code == 200
    assert b"Pre Review 1" in page.data
    assert b"Pre Review 2" in page.data
    assert b"CloudHammer" not in page.data
    assert b"GPT" not in page.data

    response = client.post(
        "/changes/change-1/review",
        data={
            "status": "pending",
            "selected_pre_review": PRE_REVIEW_2,
            "reviewer_text": "old text",
            "reviewer_notes": "",
        },
    )

    assert response.status_code == 302
    loaded = WorkspaceStore(store.workspace_dir).load()
    item = loaded.data.change_items[0]
    payload = pre_review_payload(item)
    assert payload["selected"] == PRE_REVIEW_2
    assert item.reviewer_text == "Provide new roof curb."


def test_probable_legend_item_shows_accept_as_legend_button(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    store.data.clouds[0] = replace(
        store.data.clouds[0],
        nearby_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL",
        scope_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL",
    )
    store.data.change_items[0] = replace(
        store.data.change_items[0],
        raw_text=store.data.clouds[0].scope_text,
        normalized_text=store.data.clouds[0].scope_text.lower(),
    )
    store.save()
    enrich_workspace_legend_context(store)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    page = client.get("/changes/change-1")

    assert page.status_code == 200
    assert b"Accept as legend" in page.data
    assert b"training" not in page.data.lower()
    assert b"eval" not in page.data.lower()


def test_non_probable_review_item_does_not_show_accept_as_legend(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    page = client.get("/changes/change-1")

    assert page.status_code == 200
    assert b"Accept as legend" not in page.data


def test_accept_as_legend_confirms_soft_hides_records_event_and_advances(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.data.clouds[0] = replace(
        store.data.clouds[0],
        nearby_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL",
        scope_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL",
    )
    store.data.change_items[0] = replace(
        store.data.change_items[0],
        raw_text=store.data.clouds[0].scope_text,
        normalized_text=store.data.clouds[0].scope_text.lower(),
    )
    store.save()
    enrich_workspace_legend_context(store)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.post(
        "/changes/change-1/accept-legend",
        data={"queue_status": "pending", "search_query": "", "attention_only": "0"},
        headers={"Cf-Access-Authenticated-User-Email": "reviewer@example.com"},
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/changes/change-2?queue=pending&q=")
    loaded = WorkspaceStore(store.workspace_dir).load()
    item = next(row for row in loaded.data.change_items if row.id == "change-1")
    payload = legend_context_payload(item)
    assert payload["confirmed"] is True
    assert item.superseded_reason == "legend_context"
    assert [row.id for row in visible_change_items(loaded.data.change_items)] == ["change-2"]
    assert loaded.data.review_events[-1].action == "relabel"
    assert loaded.data.review_events[-1].human_result_json["final_label"] == "legend_context"
    changes_page = client.get("/changes?status=all")
    sheet_page = client.get("/sheets/sheet-1")
    assert b'change-1' not in changes_page.data
    assert b'data-change-id="change-1"' not in sheet_page.data
    assert b'data-change-id="change-2"' in sheet_page.data


def test_exports_use_selected_pre_review_text_and_keep_multiple_boxes_one_item(tmp_path: Path):
    from openpyxl import load_workbook

    class MultiBoxProvider(FakePreReviewProvider):
        def review(self, context):
            return normalize_pre_review_2(
                {
                    "geometry_decision": "overmerged",
                    "boxes": [[45, 25, 60, 40], [115, 25, 45, 40]],
                    "refined_text": "Provide two roof curb updates.",
                    "reason": "Two visible clouded areas are inside the same review crop.",
                    "confidence": 0.77,
                    "tags": ["overmerge"],
                },
                context,
            )

    store = build_pre_review_test_store(tmp_path)
    ensure_workspace_pre_review(store, MultiBoxProvider())
    selected = select_pre_review_source(store.data.change_items[0], PRE_REVIEW_2)
    store.update_change_item(
        selected.id,
        status="approved",
        reviewer_text=selected.reviewer_text,
        provenance=selected.provenance,
    )

    outputs = Exporter(store).export(force_attention=True)
    packet = build_review_packet(store)

    assert packet.item_count == 1
    assert "Provide two roof curb updates." in packet.html_path.read_text(encoding="utf-8")
    wb = load_workbook(outputs["revision_changelog_xlsx"])
    assert "Provide two roof curb updates" in wb["Sheet1"].cell(row=2, column=5).value


def test_confirmed_legend_item_is_excluded_from_export_and_review_packet(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    item = store.data.change_items[0]
    store.data.change_items[0] = replace(
        item,
        status="approved",
        provenance={
            **item.provenance,
            LEGEND_CONTEXT_KEY: {
                "schema": "scopeledger.legend_context.v1",
                "probable": True,
                "confirmed": True,
                "symbol_definitions": [{"token": "X", "description": "REMOVE EXISTING WALL"}],
            },
        },
        superseded_reason="legend_context",
        superseded_at="2026-05-11T00:00:00+00:00",
    )
    store.save()

    outputs = Exporter(store).export(force_attention=True)
    packet = build_review_packet(store)
    approved_rows = json.loads(Path(outputs["approved_changes_json"]).read_text(encoding="utf-8"))
    pricing_rows = json.loads(Path(outputs["pricing_change_candidates_json"]).read_text(encoding="utf-8"))

    assert approved_rows == []
    assert pricing_rows == []
    assert packet.item_count == 0


def test_sheet_detail_hides_orphan_cloud_candidates(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    orphan = replace(store.data.clouds[0], id="cloud-orphan", bbox=[10, 10, 20, 20])
    store.data.clouds.append(orphan)
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.get("/sheets/sheet-1")

    assert response.status_code == 200
    assert response.data.count(b'class="bbox ') == 1
    assert b'data-change-id="change-1"' in response.data
    assert b"cloud-orphan" not in response.data


def test_sheet_detail_uses_selected_pre_review_geometry(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    ensure_workspace_pre_review(store, FakePreReviewProvider())
    selected = select_pre_review_source(store.data.change_items[0], PRE_REVIEW_2)
    store.update_change_item(selected.id, provenance=selected.provenance, reviewer_text=selected.reviewer_text)
    loaded = WorkspaceStore(store.workspace_dir).load()
    item = loaded.data.change_items[0]
    cloud = loaded.data.clouds[0]
    selected_box = selected_review_page_boxes(item, cloud)[0]
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.get("/sheets/sheet-1")

    assert response.status_code == 200
    assert f'data-x="{selected_box[0]}"'.encode("utf-8") in response.data
    assert f'data-y="{selected_box[1]}"'.encode("utf-8") in response.data
    assert f'data-w="{selected_box[2]}"'.encode("utf-8") in response.data
    assert f'data-h="{selected_box[3]}"'.encode("utf-8") in response.data
    assert b'data-x="80"' not in response.data


def test_sheet_detail_uses_crop_adjusted_geometry(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    adjusted = client.post("/changes/change-1/crop-adjustment", json={"crop_box": [50, 20, 80, 40]})
    assert adjusted.status_code == 200
    payload = adjusted.get_json()
    response = client.get("/sheets/sheet-1")

    assert response.status_code == 200
    assert f'data-x="{payload["page_box"][0]}"'.encode("utf-8") in response.data
    assert f'data-y="{payload["page_box"][1]}"'.encode("utf-8") in response.data
    assert f'data-w="{payload["page_box"][2]}"'.encode("utf-8") in response.data
    assert f'data-h="{payload["page_box"][3]}"'.encode("utf-8") in response.data


def test_sheet_detail_renders_multiple_selected_boxes_for_one_review_item(tmp_path: Path):
    class MultiBoxProvider(FakePreReviewProvider):
        def review(self, context):
            return normalize_pre_review_2(
                {
                    "geometry_decision": "overmerged",
                    "boxes": [[45, 25, 60, 40], [115, 25, 45, 40]],
                    "refined_text": "Provide two roof curb updates.",
                    "reason": "Two visible clouded areas are inside the same review crop.",
                    "confidence": 0.77,
                    "tags": ["overmerge"],
                },
                context,
            )

    store = build_pre_review_test_store(tmp_path)
    ensure_workspace_pre_review(store, MultiBoxProvider())
    selected = select_pre_review_source(store.data.change_items[0], PRE_REVIEW_2)
    store.update_change_item(selected.id, provenance=selected.provenance, reviewer_text=selected.reviewer_text)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.get("/sheets/sheet-1")

    assert response.status_code == 200
    assert response.data.count(b'class="bbox ') == 2
    assert response.data.count(b'data-change-id="change-1"') == 2
    assert response.data.count(b'class="sheet-change-card"') == 1


def test_sheet_detail_overlay_classes_follow_review_status(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    append_pre_review_test_item(store, 3)
    store.data.change_items = [
        replace(store.data.change_items[0], status="pending"),
        replace(store.data.change_items[1], status="approved"),
        replace(store.data.change_items[2], status="rejected"),
    ]
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.get("/sheets/sheet-1")

    assert response.status_code == 200
    assert b"bbox-pending" in response.data
    assert b"bbox-accepted" in response.data
    assert b"bbox-rejected" in response.data


def test_workspace_loads_without_review_events_and_preserves_events_through_rescan(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    record_review_update(
        store,
        project_id="test-project",
        change_id="change-1",
        changes={"status": "approved", "reviewer_text": "Approved scope"},
        reviewer_id="reviewer@example.com",
        review_session_id="session-1",
        action="accept",
    )
    assert len(store.data.review_events) == 1

    payload = json.loads(store.data_path.read_text(encoding="utf-8"))
    payload_without_events = dict(payload)
    payload_without_events.pop("review_events", None)
    legacy_workspace = tmp_path / "legacy-workspace"
    legacy_workspace.mkdir()
    (legacy_workspace / "workspace.json").write_text(json.dumps(payload_without_events), encoding="utf-8")
    legacy_loaded = WorkspaceStore(legacy_workspace).load()
    assert legacy_loaded.data.review_events == []

    rescanned = RevisionScanner(Path(store.data.input_dir), store.workspace_dir).scan()
    assert len(rescanned.data.review_events) == 1
    assert rescanned.data.review_events[0].action == "accept"


def test_review_route_accept_and_reject_create_review_events(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    accepted = client.post(
        "/changes/change-1/review",
        data={
            "status_override": "approved",
            "reviewer_text": "Accepted roof curb scope",
            "reviewer_notes": "Looks correct",
        },
        headers={"Cf-Access-Authenticated-User-Email": "Kevin@Example.com"},
    )

    assert accepted.status_code == 302
    loaded = WorkspaceStore(store.workspace_dir).load()
    assert len(loaded.data.review_events) == 1
    event = loaded.data.review_events[0]
    assert event.action == "accept"
    assert event.reviewer_id == "kevin@example.com"
    assert event.review_session_id
    assert event.original_candidate_json["cloud_candidate"]["bbox"] == [80, 40, 120, 80]
    assert event.human_result_json["final_text"] == "Accepted roof curb scope"
    assert event.human_result_json["final_geometry"]["boxes"]

    rejected = client.post(
        "/changes/change-1/review",
        data={
            "status_override": "rejected",
            "reviewer_text": "Accepted roof curb scope",
            "reviewer_notes": "Reject after second look",
        },
    )

    assert rejected.status_code == 302
    reloaded = WorkspaceStore(store.workspace_dir).load()
    assert [item.action for item in reloaded.data.review_events] == ["accept", "reject"]


def test_review_route_relabel_and_comment_events_stay_internal(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    ensure_workspace_pre_review(store, FakePreReviewProvider())
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    page = client.get("/changes/change-1")
    assert page.status_code == 200
    for hidden_term in (b"review event", b"truth capture", b"training data", b"eval truth"):
        assert hidden_term not in page.data.lower()

    relabel = client.post(
        "/changes/change-1/review",
        data={
            "status": "pending",
            "selected_pre_review": PRE_REVIEW_2,
            "reviewer_text": "old text",
            "reviewer_notes": "",
        },
    )
    assert relabel.status_code == 302
    loaded = WorkspaceStore(store.workspace_dir).load()
    assert loaded.data.review_events[-1].action == "relabel"
    assert loaded.data.review_events[-1].ai_suggestion_json["payload"][PRE_REVIEW_2]["text"] == "Provide new roof curb."

    comment = client.post(
        "/changes/change-1/review",
        data={
            "status": "pending",
            "reviewer_text": "Provide new roof curb.",
            "reviewer_notes": "Ask Kevin about symbol reference.",
        },
    )
    assert comment.status_code == 302
    commented = WorkspaceStore(store.workspace_dir).load()
    assert commented.data.review_events[-1].action == "comment"
    assert commented.data.review_events[-1].notes == "Ask Kevin about symbol reference."


def test_internal_review_events_support_geometry_actions(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)

    resize = record_internal_review_event(
        store,
        project_id="test-project",
        change_id="change-1",
        action="resize",
        human_result_overrides={"final_geometry": {"boxes": [[82, 42, 90, 50]]}},
    )
    merge = record_internal_review_event(
        store,
        project_id="test-project",
        change_id="change-1",
        action="merge",
        human_result_overrides={"merged_candidate_ids": ["cloud-1", "cloud-2"]},
    )
    split = record_internal_review_event(
        store,
        project_id="test-project",
        change_id="change-1",
        action="split",
        human_result_overrides={"split_child_geometries": [[[80, 40, 40, 30]], [[130, 60, 30, 30]]]},
    )

    assert resize.human_result_json["final_geometry"]["boxes"] == [[82, 42, 90, 50]]
    assert resize.original_candidate_json["cloud_candidate"]["bbox"] == [80, 40, 120, 80]
    assert merge.human_result_json["merged_candidate_ids"] == ["cloud-1", "cloud-2"]
    assert len(split.human_result_json["split_child_geometries"]) == 2


def test_crop_adjustment_coordinate_conversion():
    assert crop_box_to_page_box([50, 20, 80, 40], [40, 20, 240, 160], (200, 120)) == [
        100.0,
        46.667,
        96.0,
        53.333,
    ]


def test_crop_adjustment_route_updates_provenance_and_creates_resize_event(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    original_bbox = list(store.data.clouds[0].bbox)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    page = client.get("/changes/change-1")
    assert page.status_code == 200
    assert b"Adjust crop" in page.data
    for hidden_term in (b"training", b"labeling", b"mining", b"eval", b"CloudHammer", b"GPT"):
        assert hidden_term not in page.data

    response = client.post(
        "/changes/change-1/crop-adjustment",
        json={"crop_box": [50, 20, 80, 40]},
        headers={"Cf-Access-Authenticated-User-Email": "Kevin@Example.com"},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["crop_box"]
    loaded = WorkspaceStore(store.workspace_dir).load()
    item = loaded.data.change_items[0]
    cloud = loaded.data.clouds[0]
    adjustment = crop_adjustment_payload(item)
    assert adjustment["schema"] == CROP_ADJUSTMENT_KEY
    assert Path(adjustment["crop_image_path"]).exists()
    assert adjustment["page_boxes"] == [payload["page_box"]]
    assert selected_review_page_boxes(item, cloud) == adjustment["page_boxes"]
    assert cloud.bbox == original_bbox

    assert len(loaded.data.review_events) == 1
    event = loaded.data.review_events[0]
    assert event.action == "resize"
    assert event.reviewer_id == "kevin@example.com"
    assert event.original_candidate_json["cloud_candidate"]["bbox"] == original_bbox
    assert CROP_ADJUSTMENT_KEY not in event.original_candidate_json["change_item"]["provenance"]
    assert event.human_result_json["final_geometry"]["boxes"] == adjustment["page_boxes"]
    assert event.human_result_json["crop_adjustment"]["adjustment_id"] == adjustment["adjustment_id"]


def test_crop_adjustment_route_rejects_invalid_box_without_event(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.post("/changes/change-1/crop-adjustment", json={"crop_box": [2, 2, 3, 3]})

    assert response.status_code == 400
    loaded = WorkspaceStore(store.workspace_dir).load()
    assert loaded.data.review_events == []
    assert CROP_ADJUSTMENT_KEY not in loaded.data.change_items[0].provenance


def test_crop_adjustment_route_reports_render_failure_without_event(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    sheet = store.data.sheets[0]
    store.data.sheets[0] = SheetVersion(
        id=sheet.id,
        revision_set_id=sheet.revision_set_id,
        source_pdf=str(tmp_path / "missing.pdf"),
        page_number=sheet.page_number,
        sheet_id=sheet.sheet_id,
        sheet_title=sheet.sheet_title,
        issue_date=sheet.issue_date,
        revision_entries=sheet.revision_entries,
        narrative_entry_ids=sheet.narrative_entry_ids,
        status=sheet.status,
        render_path=sheet.render_path,
        width=sheet.width,
        height=sheet.height,
        page_text_excerpt=sheet.page_text_excerpt,
    )
    store.save()
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()

    response = client.post("/changes/change-1/crop-adjustment", json={"crop_box": [50, 20, 80, 40]})

    assert response.status_code == 400
    assert "source drawing PDF" in response.get_json()["error"]
    loaded = WorkspaceStore(store.workspace_dir).load()
    assert loaded.data.review_events == []
    assert CROP_ADJUSTMENT_KEY not in loaded.data.change_items[0].provenance


def test_exports_use_crop_adjustment_geometry_and_asset(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    register_workspace_project(store.workspace_dir)
    app = create_app(store.workspace_dir)
    client = app.test_client()
    response = client.post("/changes/change-1/crop-adjustment", json={"crop_box": [50, 20, 80, 40]})
    assert response.status_code == 200

    loaded = WorkspaceStore(store.workspace_dir).load()
    loaded.update_change_item("change-1", status="approved", reviewer_text="Adjusted crop scope")
    adjusted = WorkspaceStore(store.workspace_dir).load()
    item = adjusted.data.change_items[0]
    cloud = adjusted.data.clouds[0]
    adjustment = crop_adjustment_payload(item)

    outputs = Exporter(adjusted).export(force_attention=True)
    packet = build_review_packet(adjusted)

    assert Path(outputs["revision_changelog_xlsx"]).exists()
    assert packet.item_count == 1
    packet_asset = packet.html_path.parent / f"{packet.html_path.stem}_assets" / "0001_cloud-1_selected.png"
    assert packet_asset.exists()
    assert selected_review_page_boxes(item, cloud) == adjustment["page_boxes"]


def test_export_review_events_cli_writes_jsonl(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    record_review_update(
        store,
        project_id="test-project",
        change_id="change-1",
        changes={"status": "approved", "reviewer_text": "Approved scope"},
        reviewer_id="reviewer@example.com",
        review_session_id="session-1",
        action="accept",
    )
    output_path = tmp_path / "review_events.jsonl"

    result = cli_main(
        [
            "export-review-events",
            str(store.workspace_dir),
            "--project-id",
            "test-project",
            "--out",
            str(output_path),
        ]
    )

    assert result == 0
    rows = [json.loads(line) for line in output_path.read_text(encoding="utf-8").splitlines()]
    assert len(rows) == 1
    assert rows[0]["project_id"] == "test-project"
    assert rows[0]["action"] == "accept"
    assert rows[0]["original_candidate_json"]["cloud_candidate"]["id"] == "cloud-1"
    assert rows[0]["human_result_json"]["final_text"] == "Approved scope"


def test_enrich_workspace_scope_text_updates_existing_cloud_items(tmp_path: Path):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    input_dir.mkdir()
    pdf_path = input_dir / "Revision #1" / "drawing.pdf"
    pdf_path.parent.mkdir()
    document = fitz.open()
    page = document.new_page(width=300, height=200)
    page.insert_text((82, 95), "PROVIDE NEW GRAB BAR BLOCKING", fontsize=10)
    document.save(pdf_path)
    document.close()
    store = WorkspaceStore(workspace_dir).create(input_dir)
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf=str(pdf_path),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        width=300,
        height=200,
    )
    cloud = CloudCandidate(
        id="cloud-1",
        sheet_version_id=sheet.id,
        bbox=[70, 70, 95, 55],
        image_path="crop.png",
        page_image_path="page.png",
        confidence=0.91,
        extraction_method="cloudhammer_manifest",
        nearby_text="Cloud Only - CloudHammer detected revision cloud.",
        detail_ref=None,
        metadata={"cloudhammer_candidate_id": "cand-1"},
    )
    item = ChangeItem(
        id="change-1",
        sheet_version_id=sheet.id,
        cloud_candidate_id=cloud.id,
        sheet_id=sheet.sheet_id,
        detail_ref=None,
        raw_text=cloud.nearby_text,
        normalized_text=cloud.nearby_text.lower(),
        reviewer_text=cloud.nearby_text,
        provenance={"source": "visual-region", "extraction_method": "cloudhammer_manifest"},
    )
    store.data.sheets = [sheet]
    store.data.clouds = [cloud]
    store.data.change_items = [item]
    store.save()

    changed = enrich_workspace_scope_text(store)

    assert changed == 1
    assert store.data.clouds[0].scope_reason == "text-layer-near-cloud"
    assert "PROVIDE NEW GRAB BAR BLOCKING" in store.data.change_items[0].raw_text
    assert store.data.change_items[0].reviewer_text == store.data.change_items[0].raw_text
    assert store.data.change_items[0].provenance["cloudhammer_candidate_id"] == "cand-1"


def test_legend_context_classifier_extracts_symbol_definitions():
    text = "DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL\nZ.9 PROVIDE ROOF INFILL"

    definitions = extract_symbol_definitions(text)
    classification = classify_legend_context(text)

    assert classification.probable is True
    assert {definition["token"] for definition in definitions} == {"X", "Z.9"}
    assert any(definition["description"] == "REMOVE EXISTING WALL" for definition in definitions)


def test_legend_context_classifier_does_not_flag_text_heavy_scope():
    text = (
        "Cloud Only - PROVIDE EXHAUST ANTENNA AND CABLING GUARD RAIL, "
        "PAINT NEW BOOT FOR EXISTING VENT AND EXTEND LAUNDRY CHUTE THROUGH ROOF."
    )

    classification = classify_legend_context(text)

    assert classification.probable is False


def test_legend_context_enrichment_flags_probable_and_resolves_same_sheet(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    legend_text = "DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL\nZ.9 PROVIDE ROOF INFILL"
    scope_text = "Cloud Only - X"
    store.data.clouds[0] = replace(store.data.clouds[0], nearby_text=legend_text, scope_text=legend_text)
    store.data.change_items[0] = replace(
        store.data.change_items[0],
        raw_text=legend_text,
        normalized_text=legend_text.lower(),
    )
    store.data.clouds[1] = replace(store.data.clouds[1], nearby_text=scope_text, scope_text=scope_text)
    store.data.change_items[1] = replace(
        store.data.change_items[1],
        raw_text=scope_text,
        normalized_text=scope_text.lower(),
    )
    store.save()

    changed = enrich_workspace_legend_context(store)

    assert changed == 2
    legend_payload = legend_context_payload(store.data.change_items[0])
    scope_payload = legend_context_payload(store.data.change_items[1])
    assert legend_payload["probable"] is True
    assert legend_payload["confirmed"] is False
    assert any(definition["token"] == "X" for definition in legend_payload["symbol_definitions"])
    assert scope_payload["resolved_references"] == [
        {
            "token": "X",
            "description": "REMOVE EXISTING WALL",
            "legend_item_id": "change-1",
            "source": "same_sheet",
        }
    ]


def test_legend_context_ambiguous_package_fallback_is_not_attached(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    base_sheet = store.data.sheets[0]
    base_cloud = store.data.clouds[0]
    base_item = store.data.change_items[0]
    sheet_2 = replace(base_sheet, id="sheet-2", sheet_id="AE102")
    sheet_3 = replace(base_sheet, id="sheet-3", sheet_id="AE103")
    cloud_2 = replace(base_cloud, id="cloud-2", sheet_version_id=sheet_2.id)
    cloud_3 = replace(base_cloud, id="cloud-3", sheet_version_id=sheet_3.id, nearby_text="Cloud Only - X", scope_text="Cloud Only - X")
    store.data.sheets.extend([sheet_2, sheet_3])
    store.data.clouds = [
        replace(base_cloud, nearby_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL", scope_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL"),
        replace(cloud_2, nearby_text="DEMOLITION PLAN LEGEND\nX PATCH FLOOR OPENING", scope_text="DEMOLITION PLAN LEGEND\nX PATCH FLOOR OPENING"),
        cloud_3,
    ]
    store.data.change_items = [
        replace(base_item, raw_text=store.data.clouds[0].scope_text, normalized_text=store.data.clouds[0].scope_text.lower()),
        replace(
            base_item,
            id="change-2",
            sheet_version_id=sheet_2.id,
            cloud_candidate_id=cloud_2.id,
            sheet_id=sheet_2.sheet_id,
            raw_text=store.data.clouds[1].scope_text,
            normalized_text=store.data.clouds[1].scope_text.lower(),
        ),
        replace(
            base_item,
            id="change-3",
            sheet_version_id=sheet_3.id,
            cloud_candidate_id=cloud_3.id,
            sheet_id=sheet_3.sheet_id,
            raw_text="Cloud Only - X",
            normalized_text="cloud only - x",
        ),
    ]
    store.save()

    enrich_workspace_legend_context(store)

    assert legend_context_payload(store.data.change_items[2]) == {}


def test_superseded_unconfirmed_probable_legend_is_not_used_as_context(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    legend_text = "DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL"
    store.data.clouds[0] = replace(store.data.clouds[0], nearby_text=legend_text, scope_text=legend_text)
    store.data.change_items[0] = replace(
        store.data.change_items[0],
        raw_text=legend_text,
        normalized_text=legend_text.lower(),
        superseded_reason="overmerge_split",
        superseded_at="2026-05-11T00:00:00+00:00",
    )
    store.data.clouds[1] = replace(store.data.clouds[1], nearby_text="Cloud Only - X", scope_text="Cloud Only - X")
    store.data.change_items[1] = replace(
        store.data.change_items[1],
        raw_text="Cloud Only - X",
        normalized_text="cloud only - x",
    )
    store.save()

    changed = enrich_workspace_legend_context(store)

    assert changed == 0
    assert legend_context_payload(store.data.change_items[0]) == {}
    assert legend_context_payload(store.data.change_items[1]) == {}


def test_pre_review_one_carries_resolved_legend_context(tmp_path: Path):
    store = build_pre_review_test_store(tmp_path)
    append_pre_review_test_item(store, 2)
    store.data.clouds[0] = replace(store.data.clouds[0], nearby_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL", scope_text="DEMOLITION PLAN LEGEND\nX REMOVE EXISTING WALL")
    store.data.change_items[0] = replace(store.data.change_items[0], raw_text=store.data.clouds[0].scope_text, normalized_text=store.data.clouds[0].scope_text.lower())
    store.data.clouds[1] = replace(store.data.clouds[1], nearby_text="Cloud Only - X", scope_text="Cloud Only - X")
    store.data.change_items[1] = replace(store.data.change_items[1], raw_text="Cloud Only - X", normalized_text="cloud only - x")
    store.save()
    enrich_workspace_legend_context(store)

    ensure_workspace_pre_review(store)

    payload = pre_review_payload(WorkspaceStore(store.workspace_dir).load().data.change_items[1])
    assert payload[PRE_REVIEW_1]["text"] == "Cloud Only - X"
    assert payload[PRE_REVIEW_1]["legend_context"] == "X: REMOVE EXISTING WALL"


def test_export_refreshes_ocr_scope_text_before_workbook(tmp_path: Path):
    from openpyxl import load_workbook
    from PIL import Image

    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    pdf_path = input_dir / "Revision #1 - Drawing Changes" / "drawing.pdf"
    pdf_path.parent.mkdir(parents=True)
    document = fitz.open()
    page = document.new_page(width=300, height=200)
    page.insert_text((82, 95), "PROVIDE NEW GRAB BAR BLOCKING", fontsize=10)
    document.save(pdf_path)
    document.close()

    store = WorkspaceStore(workspace_dir).create(input_dir)
    render_path = store.page_path("sheet-1")
    crop_path = store.crop_path("cloud-1")
    Image.new("RGB", (300, 200), "white").save(render_path)
    Image.new("RGB", (120, 80), "white").save(crop_path)
    store.data.revision_sets = [
        RevisionSet(
            id="rev-1",
            label="Revision #1 - Drawing Changes",
            source_dir=str(pdf_path.parent),
            set_number=1,
            set_date="04/28/2026",
            pdf_paths=[str(pdf_path)],
        )
    ]
    store.data.sheets = [
        SheetVersion(
            id="sheet-1",
            revision_set_id="rev-1",
            source_pdf=str(pdf_path),
            page_number=1,
            sheet_id="AE101",
            sheet_title="Preview",
            issue_date="04/28/2026",
            render_path=str(render_path),
            width=300,
            height=200,
            status="active",
        )
    ]
    placeholder = "Cloud Only - CloudHammer detected revision cloud. OCR/scope extraction is not wired yet."
    store.data.clouds = [
        CloudCandidate(
            id="cloud-1",
            sheet_version_id="sheet-1",
            bbox=[70, 70, 95, 55],
            image_path=str(crop_path),
            page_image_path=str(render_path),
            confidence=0.91,
            extraction_method="cloudhammer_manifest",
            nearby_text=placeholder,
            detail_ref=None,
        )
    ]
    store.data.change_items = [
        ChangeItem(
            id="change-1",
            sheet_version_id="sheet-1",
            cloud_candidate_id="cloud-1",
            sheet_id="AE101",
            detail_ref=None,
            raw_text=placeholder,
            normalized_text=placeholder.lower(),
            provenance={"source": "visual-region", "extraction_method": "cloudhammer_manifest"},
            status="approved",
            reviewer_text=placeholder,
        )
    ]
    store.save()

    exporter = Exporter(store)
    outputs = exporter.export(force_attention=True)

    assert exporter.last_scope_enrichment_count >= 1
    assert "PROVIDE NEW GRAB BAR BLOCKING" in store.data.change_items[0].raw_text
    approved = json.loads((store.output_dir / "approved_changes.json").read_text(encoding="utf-8"))
    assert "PROVIDE NEW GRAB BAR BLOCKING" in approved[0]["text"]
    wb = load_workbook(outputs["revision_changelog_xlsx"])
    assert "PROVIDE NEW GRAB BAR BLOCKING" in wb["Sheet1"].cell(row=2, column=5).value
    assert "Review Flags" in wb.sheetnames
    flags = wb["Review Flags"]
    assert "PROVIDE NEW GRAB BAR BLOCKING" in flags.cell(row=2, column=15).value
    assert flags.cell(row=2, column=12).value == "text-layer-near-cloud"


def test_cloudhammer_manifest_clouds_get_visual_items_even_with_narratives():
    scanner = RevisionScanner.__new__(RevisionScanner)
    narrative = NarrativeEntry(
        id="narrative-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        heading="Revision narrative",
        summary="Replace existing fixture.",
    )
    sheet = SheetVersion(
        id="sheet-1",
        revision_set_id="rev-1",
        source_pdf="revision.pdf",
        page_number=1,
        sheet_id="AE101",
        sheet_title="Preview",
        issue_date=None,
        narrative_entry_ids=[narrative.id],
    )
    clouds = [
        CloudCandidate(
            id="cloud-1",
            sheet_version_id=sheet.id,
            bbox=[0, 0, 10, 10],
            image_path="crop-1.png",
            page_image_path="page.png",
            confidence=0.9,
            extraction_method="cloudhammer_manifest",
            nearby_text="Cloud Only - CloudHammer detected revision cloud. candidate=c1",
            detail_ref=None,
        ),
        CloudCandidate(
            id="cloud-2",
            sheet_version_id=sheet.id,
            bbox=[20, 20, 10, 10],
            image_path="crop-2.png",
            page_image_path="page.png",
            confidence=0.8,
            extraction_method="cloudhammer_manifest",
            nearby_text="Cloud Only - CloudHammer detected revision cloud. candidate=c2",
            detail_ref=None,
        ),
    ]

    items = scanner._generate_change_items([narrative], [sheet], clouds)

    assert len(items) == 3
    assert Counter(item.provenance["source"] for item in items) == {"narrative": 1, "visual-region": 2}
    assert {item.cloud_candidate_id for item in items if item.provenance["source"] == "visual-region"} == {"cloud-1", "cloud-2"}


def test_review_state_restores_by_cloud_id_when_extracted_text_changes():
    scanner = RevisionScanner.__new__(RevisionScanner)
    scanner.previous_change_items = [
        ChangeItem(
            id="old",
            sheet_version_id="sheet-1",
            cloud_candidate_id="cloud-1",
            sheet_id="AE101",
            detail_ref=None,
            raw_text="CloudHammer detected revision cloud",
            normalized_text="cloudhammer detected revision cloud",
            status="approved",
            reviewer_text="Preserved reviewer edit",
        )
    ]
    new_item = ChangeItem(
        id="new",
        sheet_version_id="sheet-1",
        cloud_candidate_id="cloud-1",
        sheet_id="AE101",
        detail_ref=None,
        raw_text="Cloud Only - PROVIDE NEW GRAB BAR BLOCKING",
        normalized_text="cloud only - provide new grab bar blocking",
    )

    restored = scanner._restore_review_state([new_item])[0]

    assert restored.status == "approved"
    assert restored.reviewer_text == "Preserved reviewer edit"
    assert restored.raw_text == new_item.raw_text


def test_export_only_approved_items_when_forced(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    pending = next(item for item in store.data.change_items if item.status == "pending")
    store.update_change_item(pending.id, status="approved", reviewer_text="Verified approved scope")

    exporter = Exporter(store)
    outputs = exporter.export(force_attention=True)
    rows = json.loads((store.output_dir / "approved_changes.json").read_text(encoding="utf-8"))
    pricing_candidates = json.loads((store.output_dir / "pricing_change_candidates.json").read_text(encoding="utf-8"))
    pricing_log = json.loads((store.output_dir / "pricing_change_log.json").read_text(encoding="utf-8"))
    conformed_index = json.loads((store.output_dir / "conformed_sheet_index.json").read_text(encoding="utf-8"))
    diagnostics = json.loads((store.output_dir / "preflight_diagnostics.json").read_text(encoding="utf-8"))

    assert rows
    assert rows[0]["text"] == "Verified approved scope"
    assert pricing_candidates
    assert any(row["change_id"] == pending.id for row in pricing_candidates)
    assert pricing_log
    assert pricing_log[0]["pricing_status"] == "approved"
    assert pricing_log[0]["change_summary"] == "Verified approved scope"
    assert conformed_index
    assert any(row["latest_for_pricing"] for row in conformed_index)
    assert "conformed_preview_pdf" in outputs
    assert "pricing_change_candidates_csv" in outputs
    assert "pricing_change_log_csv" in outputs
    assert "conformed_sheet_index_csv" in outputs
    assert "preflight_diagnostics_csv" in outputs
    assert diagnostics["issues"]
    assert store.data.exports[-1]["forced_attention"] is True

    summary = exporter.last_summary
    assert summary["pricing_log_count"] == len(pricing_log)
    assert summary["pricing_candidate_count"] == len(pricing_candidates)
    assert summary["active_sheet_count"] == len([sheet for sheet in store.data.sheets if sheet.status == "active"])
    assert summary["superseded_sheet_count"] == len([sheet for sheet in store.data.sheets if sheet.status == "superseded"])
    assert summary["revision_set_count"] == len(store.data.revision_sets)
    assert store.data.exports[-1]["summary"] == summary


def test_revision_changelog_xlsx_matches_expected_layout(workspace_copy):
    """Smoke test that the revision changelog exporter preserves workbook shape."""
    from openpyxl import load_workbook

    from backend.deliverables.revision_changelog_excel import COLUMNS, ROWS_PER_GROUP

    store = WorkspaceStore(workspace_copy).load()
    first_pending, second_pending = store.data.change_items[:2]
    store.update_change_item(first_pending.id, status="approved", reviewer_text="Install gypsum board and corner bead")
    store.update_change_item(second_pending.id, status="approved", reviewer_text="Patch and repair masonry opening")

    outputs = Exporter(store).export(force_attention=True)
    xlsx_path = Path(outputs["revision_changelog_xlsx"])
    assert xlsx_path.exists(), "Revision changelog workbook should be written"
    assert xlsx_path.suffix == ".xlsx"

    wb = load_workbook(xlsx_path)
    assert wb.sheetnames[:2] == ["Summary", "Sheet1"]
    assert "Review Flags" in wb.sheetnames
    summary_ws = wb["Summary"]
    assert summary_ws["A1"].value == "ScopeLedger Revision Review"
    assert summary_ws["B8"].value == "Revision Sets"
    assert summary_ws["B9"].value == len(store.data.revision_sets)
    ws = wb["Sheet1"]

    headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
    expected_headers = [header or None for header, _ in COLUMNS]
    assert headers == expected_headers, "Headers must mirror Kevin's mod_5_changelog.xlsx exactly (typo and trailing spaces preserved)"
    assert headers[9] == "Qoute Received?", "Preserve Kevin's typo verbatim until he renames the column"
    assert ws.freeze_panes == "A2"
    assert ws.sheet_view.showGridLines is False
    assert ws.cell(row=1, column=1).fill.fgColor.rgb.endswith("111827")

    drawing_col_values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    populated_drawings = [v for v in drawing_col_values if v]
    assert populated_drawings, "Should have at least one row of approved data"
    for value in populated_drawings:
        assert "-" in value, f"Drawing column should be hyphenated like AE-110, got {value!r}"

    correlations = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    for corr in correlations:
        assert "." in corr, f"Correlation should be <sheet>.<seq>, got {corr!r}"

    detail_values = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=4).value]
    assert any("Cloud Only" in v or "Detail" in v for v in detail_values)

    assert ws.row_dimensions[2].height >= 16
    populated_row_count = sum(1 for v in drawing_col_values if v)
    assert (ws.max_row - 1) >= populated_row_count * ROWS_PER_GROUP - 1, "Each entry should reserve a vertical block for the embedded crop"

    assert ws.merged_cells.ranges, "Scope and Detail View columns should be merged within each block"
    assert not ws._images, "No embedded crop images are expected while CloudHammer inference is disconnected"
    review_ws = wb["Review Flags"]
    review_headers = [review_ws.cell(row=1, column=i).value for i in range(1, 16)]
    assert review_headers[:4] == ["Change ID", "Status", "Needs Review", "Review Reason"]
    assert review_ws.freeze_panes == "A2"


def test_crop_comparison_uses_previous_sheet_area(tmp_path: Path):
    from PIL import Image, ImageStat

    from backend.deliverables.crop_comparison import build_cloud_comparison_image, find_previous_sheet_version

    def write_color_pdf(path: Path, fill: tuple[float, float, float]) -> None:
        document = fitz.open()
        page = document.new_page(width=300, height=200)
        page.draw_rect(page.rect, color=fill, fill=fill)
        document.save(path)
        document.close()

    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    input_dir.mkdir()
    previous_pdf = input_dir / "previous.pdf"
    current_pdf = input_dir / "current.pdf"
    write_color_pdf(previous_pdf, (1, 0, 0))
    write_color_pdf(current_pdf, (0, 0, 1))

    store = WorkspaceStore(workspace_dir).create(input_dir)
    previous_sheet = SheetVersion(
        id="sheet-prev",
        revision_set_id="rev-1",
        source_pdf=str(previous_pdf),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Plan",
        issue_date="04/01/2026",
        width=300,
        height=200,
    )
    current_sheet = SheetVersion(
        id="sheet-current",
        revision_set_id="rev-2",
        source_pdf=str(current_pdf),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Plan",
        issue_date="04/15/2026",
        width=300,
        height=200,
    )
    store.data.revision_sets = [
        RevisionSet(id="rev-1", label="Revision #1", source_dir=str(input_dir), set_number=1, set_date="04/01/2026"),
        RevisionSet(id="rev-2", label="Revision #2", source_dir=str(input_dir), set_number=2, set_date="04/15/2026"),
    ]
    store.data.sheets = [previous_sheet, current_sheet]
    cloud = CloudCandidate(
        id="cloud-1",
        sheet_version_id=current_sheet.id,
        bbox=[80, 60, 80, 60],
        image_path="",
        page_image_path="",
        confidence=0.9,
        extraction_method="cloudhammer_manifest",
        nearby_text="",
        detail_ref=None,
    )

    previous = find_previous_sheet_version(
        current_sheet,
        store.data.sheets,
        {revision_set.id: revision_set for revision_set in store.data.revision_sets},
    )
    output = build_cloud_comparison_image(
        store,
        cloud=cloud,
        current_sheet=current_sheet,
        previous_sheet=previous,
        output_path=tmp_path / "comparison.png",
    )

    assert output and output.exists()
    with Image.open(output) as image:
        left_mean = ImageStat.Stat(image.crop((20, 60, image.width // 2 - 20, image.height - 20))).mean
        right_mean = ImageStat.Stat(image.crop((image.width // 2 + 20, 60, image.width - 20, image.height - 20))).mean
    assert left_mean[0] > left_mean[2], "previous panel should use the red previous PDF"
    assert right_mean[2] > right_mean[0], "current panel should use the blue current PDF"


def test_crop_comparison_requires_prior_real_revision_set():
    from backend.deliverables.crop_comparison import find_previous_sheet_version

    current_rev1 = SheetVersion(
        id="current-rev1",
        revision_set_id="rev-1",
        source_pdf="current.pdf",
        page_number=4,
        sheet_id="AE102",
        sheet_title="Second Floor Plan",
        issue_date="10/10/2025",
    )
    same_package_index = SheetVersion(
        id="index-rev1",
        revision_set_id="rev-1",
        source_pdf="current.pdf",
        page_number=1,
        sheet_id="AE102",
        sheet_title="SHEET INDEX - CONFORMED SET PAGE NO. SHEET NO. SHEET NAME AE102 2ND FLOOR PLAN X",
        issue_date="10/10/2025",
    )
    previous_real = SheetVersion(
        id="real-rev1",
        revision_set_id="rev-1",
        source_pdf="previous.pdf",
        page_number=4,
        sheet_id="AE102",
        sheet_title="Second Floor Plan",
        issue_date="08/15/2025",
    )
    current_rev2 = SheetVersion(
        id="current-rev2",
        revision_set_id="rev-2",
        source_pdf="current-rev2.pdf",
        page_number=4,
        sheet_id="AE102",
        sheet_title="Second Floor Plan",
        issue_date="10/10/2025",
    )
    revision_sets = {
        "rev-1": RevisionSet(id="rev-1", label="Revision #1", source_dir="", set_number=1, set_date="10/10/2025"),
        "rev-2": RevisionSet(id="rev-2", label="Revision #2", source_dir="", set_number=2, set_date="10/17/2025"),
    }

    assert find_previous_sheet_version(current_rev1, [same_package_index, current_rev1], revision_sets) is None

    previous = find_previous_sheet_version(current_rev2, [same_package_index, previous_real, current_rev2], revision_sets)

    assert previous == previous_real


def test_scanner_treats_revision_set_upload_as_rev1_and_skips_index_clouds(tmp_path: Path):
    class FakeCloudClient:
        name = "fake_clouds"
        cache_key = "fake_clouds"

        def detect(self, *, page, sheet):
            return [
                CloudDetection(
                    bbox=[60, 60, 120, 80],
                    confidence=0.95,
                    extraction_method="cloudhammer_manifest",
                    nearby_text="Detected revision region",
                )
            ]

    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    pdf_path = input_dir / "Revision Set 1" / "package.pdf"
    pdf_path.parent.mkdir(parents=True)
    document = fitz.open()
    index_page = document.new_page(width=600, height=800)
    index_page.insert_text((60, 80), "SHEET INDEX - CONFORMED SET PAGE NO. SHEET NO. SHEET NAME AE102 2ND FLOOR PLAN X", fontsize=10)
    index_page.insert_text((430, 720), "AE102", fontsize=14)
    real_page = document.new_page(width=600, height=800)
    real_page.insert_text((60, 80), "AE102 SECOND FLOOR PLAN PROVIDE NEW WALL PATCHING", fontsize=12)
    real_page.insert_text((430, 720), "AE102", fontsize=14)
    real_page.insert_text((430, 742), "SECOND FLOOR PLAN", fontsize=10)
    document.save(pdf_path)
    document.close()

    store = RevisionScanner(input_dir, workspace_dir, cloud_inference_client=FakeCloudClient()).scan()

    assert store.data.revision_sets[0].label == "Revision Set 1"
    assert store.data.revision_sets[0].set_number == 1
    assert len(store.data.sheets) == 2
    index_sheet = next(sheet for sheet in store.data.sheets if sheet_is_index_like(sheet))
    real_sheet = next(sheet for sheet in store.data.sheets if sheet.id != index_sheet.id)
    assert index_sheet.status == "superseded"
    assert real_sheet.status == "active"
    assert len(store.data.clouds) == 1
    assert store.data.clouds[0].sheet_version_id == real_sheet.id
    assert {item.sheet_version_id for item in store.data.change_items} == {real_sheet.id}


def test_revision_changelog_stacks_multiple_items_in_same_cloud(tmp_path: Path):
    """Kevin confirmed same-cloud scope items can share one row if listed."""
    from openpyxl import load_workbook
    from PIL import Image

    from backend.deliverables.revision_changelog_excel import ROWS_PER_GROUP, write_revision_changelog

    store = WorkspaceStore(tmp_path / "workspace").create(tmp_path / "input")
    crop_path = tmp_path / "same_cloud.png"
    Image.new("RGB", (180, 100), "white").save(crop_path)
    store.data.revision_sets = [
        RevisionSet(
            id="rev-1",
            label="Revision #1 - Drawing Changes",
            source_dir=str(tmp_path),
            set_number=1,
            set_date="04/28/2026",
        )
    ]
    store.data.sheets = [
        SheetVersion(
            id="sheet-1",
            revision_set_id="rev-1",
            source_pdf=str(tmp_path / "revision.pdf"),
            page_number=1,
            sheet_id="AE101",
            sheet_title="Plan",
            issue_date="04/28/2026",
        )
    ]
    store.data.clouds = [
        CloudCandidate(
            id="cloud-1",
            sheet_version_id="sheet-1",
            bbox=[10, 20, 180, 100],
            image_path=str(crop_path),
            page_image_path="",
            confidence=0.9,
            extraction_method="test",
            nearby_text="",
            detail_ref=None,
        )
    ]
    store.data.change_items = [
        ChangeItem(
            id="item-1",
            sheet_version_id="sheet-1",
            cloud_candidate_id="cloud-1",
            sheet_id="AE101",
            detail_ref=None,
            raw_text="Install grab bar blocking",
            normalized_text="install grab bar blocking",
            status="approved",
            reviewer_text="Install grab bar blocking",
        ),
        ChangeItem(
            id="item-2",
            sheet_version_id="sheet-1",
            cloud_candidate_id="cloud-1",
            sheet_id="AE101",
            detail_ref=None,
            raw_text="Patch wall tile",
            normalized_text="patch wall tile",
            status="approved",
            reviewer_text="Patch wall tile",
        ),
    ]

    output_path = write_revision_changelog(store, tmp_path / "revision_changelog.xlsx")

    wb = load_workbook(output_path)
    assert wb.sheetnames[:2] == ["Summary", "Sheet1"]
    assert wb["Summary"]["B9"].value == 1
    ws = wb["Sheet1"]
    populated_drawings = [
        ws.cell(row=row, column=2).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=2).value
    ]
    assert populated_drawings == ["AE-101"]
    assert ws.max_row == 1 + ROWS_PER_GROUP
    assert ws.cell(row=2, column=4).value == "N/A - Cloud Only "
    assert ws.cell(row=2, column=5).value == "1) Install grab bar blocking\n2) Patch wall tile"
    assert len(ws._images) == 1


def test_pricing_outputs_filter_placeholder_revision_regions(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    seed_item = next(item for item in store.data.change_items if item.status == "pending")
    placeholder_text = (
        f"Possible revision region near {seed_item.detail_ref}"
        if seed_item.detail_ref
        else "Possible revision region"
    )
    store.update_change_item(seed_item.id, status="approved", reviewer_text=placeholder_text)
    real_item = next(
        item
        for item in store.data.change_items
        if item.id != seed_item.id and item.status == "pending"
    )
    store.update_change_item(
        real_item.id,
        status="approved",
        reviewer_text="Patch and repair gypsum board, install new corner bead",
    )

    exporter = Exporter(store)
    exporter.export(force_attention=True)
    candidates = json.loads((store.output_dir / "pricing_change_candidates.json").read_text(encoding="utf-8"))
    log = json.loads((store.output_dir / "pricing_change_log.json").read_text(encoding="utf-8"))

    assert all(row["change_id"] != seed_item.id for row in candidates), "Placeholder rows should be filtered from candidates"
    assert all(row["change_id"] != seed_item.id for row in log), "Placeholder rows should never appear in the pricing log"
    assert any(row["change_id"] == real_item.id for row in log), "Real approved scope should still reach the pricing log"

    filtered = exporter.last_summary["filtered_by_reason"]
    assert filtered.get("placeholder-no-readable-scope", 0) >= 1


def test_pricing_relevance_filter_drops_label_locator_combos():
    exporter = Exporter.__new__(Exporter)
    cases = [
        ("AE113 First Floor Plan", ["ROOM", "RADIO"], "ROOM; RADIO", "visual-region", "pending"),
        ("AE125 Plan", ["AE600", "SEE DETAIL FOR"], "AE600; SEE DETAIL FOR", "visual-region", "pending"),
        ("AE125 Plan", ["WOMENS", "AE403"], "WOMENS; AE403", "visual-region", "pending"),
        ("AE125 Plan", ["FLOOR", "AE514"], "FLOOR; AE514", "visual-region", "pending"),
        ("AE125 Plan", ["SCHED", "GLAZING"], "SCHED; GLAZING", "visual-region", "pending"),
        ("AE125 Plan", ["5A127", "F-1"], "5A127; F-1", "visual-region", "pending"),
        ("AE125 Plan", ["5S/5NB-38"], "5S/5NB-38", "visual-region", "pending"),
    ]
    for title, scope_lines, combined, source_kind, status in cases:
        reason = exporter._pricing_relevance_reason(title, scope_lines, combined, source_kind, status)
        assert reason == "locator-only-text", f"expected locator filter for {scope_lines!r}, got {reason!r}"


def test_pricing_relevance_filter_keeps_real_pricing_scope():
    exporter = Exporter.__new__(Exporter)
    keep_cases = [
        ("AE113 Plan", ["FIRE CAULK ALL", "2 1/2\" MTL STUDS"], "FIRE CAULK ALL; 2 1/2\" MTL STUDS"),
        ("AE401 Plan", ["EXISTING", "CONC COLUMN"], "EXISTING; CONC COLUMN"),
        ("MP105 Plan", ["4\" CHWS", "SEE DETAIL"], "4\" CHWS; SEE DETAIL"),
        ("AE600 Plan", ["PLAN", "PARTITION PER CODE"], "PLAN; PARTITION PER CODE"),
        ("MP102 Plan", ["1\" HWR", "1\" HWS"], "1\" HWR; 1\" HWS"),
        ("MH104 Plan", ["HEPA FILTER RACK", "26X22 DOWN"], "HEPA FILTER RACK; 26X22 DOWN"),
    ]
    for title, scope_lines, combined in keep_cases:
        reason = exporter._pricing_relevance_reason(title, scope_lines, combined, "visual-region", "pending")
        assert reason == "likely-pricing-scope", f"expected keep for {scope_lines!r}, got {reason!r}"


def test_pricing_relevance_filter_drops_low_signal_text_without_scope_keywords():
    exporter = Exporter.__new__(Exporter)
    reason = exporter._pricing_relevance_reason(
        "AE125 Plan",
        ["6TH FLOOR 66'-0\""],
        "6TH FLOOR 66'-0\"",
        "visual-region",
        "pending",
    )
    assert reason in ("low-signal-no-scope-keyword", "locator-only-text")


def test_pricing_relevance_filter_trusts_reviewer_for_approved_items():
    exporter = Exporter.__new__(Exporter)
    reason = exporter._pricing_relevance_reason(
        "AE125 Plan",
        ["Approved scope per RFI"],
        "Approved scope per RFI",
        "visual-region",
        "approved",
    )
    assert reason == "reviewer-approved"


def test_pricing_relevance_filter_overrides_reviewer_for_placeholder_text():
    exporter = Exporter.__new__(Exporter)
    reason = exporter._pricing_relevance_reason(
        "AE113 Plan",
        ["Possible revision region near 3/AE113"],
        "Possible revision region near 3/AE113",
        "visual-region",
        "approved",
    )
    assert reason == "placeholder-no-readable-scope"


def test_cli_export_summary_is_human_readable():
    from backend.cli import format_export_summary

    summary = {
        "output_dir": "/tmp/workspace/outputs",
        "approved_count": 4,
        "pending_count": 2,
        "rejected_count": 1,
        "attention_pending_count": 1,
        "force_attention": True,
        "pricing_log_count": 3,
        "pricing_candidate_count": 5,
        "filtered_count": 9,
        "filtered_by_reason": {
            "placeholder-no-readable-scope": 6,
            "sheet-index-page": 3,
        },
        "active_sheet_count": 12,
        "superseded_sheet_count": 4,
        "revision_set_count": 2,
    }
    outputs = {
        "pricing_change_log_csv": "/tmp/workspace/outputs/pricing_change_log.csv",
        "pricing_change_candidates_csv": "/tmp/workspace/outputs/pricing_change_candidates.csv",
        "conformed_sheet_index_csv": "/tmp/workspace/outputs/conformed_sheet_index.csv",
        "conformed_preview_pdf": "/tmp/workspace/outputs/conformed_preview.pdf",
    }
    text = format_export_summary(summary, outputs, Path("/tmp/workspace"))

    assert "Pricing-ready rows" in text
    assert "3 items" in text
    assert "pricing_change_log.csv" in text
    assert "Conformed sheet set" in text
    assert "12 latest sheets" in text
    assert "Filtered out as noise" in text
    assert "clouded regions with no readable text" in text
    assert "WARNING" in text
    assert "python -m backend serve" in text


def test_empty_project_registry_starts_without_demo(tmp_path: Path):
    app = create_app(tmp_path / "app-data")
    client = app.test_client()

    projects = client.get("/projects")
    assert projects.status_code == 200
    assert b"Demo Project" not in projects.data
    assert b"No active projects." in projects.data
    assert b"No project selected" in projects.data

    overview = client.get("/overview")
    assert overview.status_code == 200
    assert b"No project selected" in overview.data
    assert b"No revision packages have been loaded." in overview.data
    assert b"Choose PDF(s)" not in overview.data

    for path, expected in [
        ("/sheets", b"No drawings match the current filters."),
        ("/changes", b"No changes match the current filters."),
        ("/conformed", b"No sheets match the current filter."),
        ("/export", b"No project workspace is selected."),
        ("/diagnostics", b"No PDFs have been scanned."),
        ("/settings", b"No historical review-assist records were found"),
    ]:
        response = client.get(path)
        assert response.status_code == 200
        assert expected in response.data

    populate = client.post("/workspace/populate")
    assert populate.status_code == 302
    assert populate.headers["Location"].endswith("/projects")


def test_create_app_loads_allowlisted_values_from_cloudhammer_env(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    monkeypatch.delenv("SCOPELEDGER_CLOUDHAMMER_MODEL", raising=False)
    monkeypatch.delenv("SCOPELEDGER_CLOUDHAMMER_TIMEOUT_SECONDS", raising=False)
    monkeypatch.delenv("SCOPELEDGER_PREREVIEW_ENABLED", raising=False)
    monkeypatch.delenv("SCOPELEDGER_PREREVIEW_MODEL", raising=False)

    env_dir = tmp_path / "CloudHammer"
    env_dir.mkdir()
    env_path = env_dir / ".env"
    env_path.write_text(
        "\n".join(
            [
                "OPENAI_API_KEY=test-local-key",
                "SCOPELEDGER_CLOUDHAMMER_MODEL=CloudHammer/runs/test/weights/best.pt",
                "SCOPELEDGER_CLOUDHAMMER_TIMEOUT_SECONDS=120",
                "SCOPELEDGER_PREREVIEW_ENABLED=1",
                'SCOPELEDGER_PREREVIEW_MODEL="test-model"',
                "UNRELATED_SECRET=do-not-load",
            ]
        ),
        encoding="utf-8",
    )

    app = create_app(tmp_path / "app-data")

    assert os.environ["OPENAI_API_KEY"] == "test-local-key"
    assert os.environ["SCOPELEDGER_CLOUDHAMMER_MODEL"] == "CloudHammer/runs/test/weights/best.pt"
    assert os.environ["SCOPELEDGER_CLOUDHAMMER_TIMEOUT_SECONDS"] == "120"
    assert os.environ["SCOPELEDGER_PREREVIEW_ENABLED"] == "1"
    assert os.environ["SCOPELEDGER_PREREVIEW_MODEL"] == "test-model"
    assert "UNRELATED_SECRET" not in os.environ
    assert app.config["SCOPELEDGER_LOADED_ENV_FILES"] == (env_path.resolve(),)
    provider = app.config["PRE_REVIEW_PROVIDER"]
    assert isinstance(provider, OpenAIPreReviewProvider)
    assert provider.enabled is True
    assert provider.model == "test-model"


def test_create_app_does_not_override_existing_env_values(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("OPENAI_API_KEY", "already-set-key")
    monkeypatch.delenv("SCOPELEDGER_PREREVIEW_ENABLED", raising=False)

    env_dir = tmp_path / "CloudHammer"
    env_dir.mkdir()
    env_path = env_dir / ".env"
    env_path.write_text(
        "\n".join(
            [
                "OPENAI_API_KEY=file-key",
                "SCOPELEDGER_PREREVIEW_ENABLED=1",
            ]
        ),
        encoding="utf-8",
    )

    create_app(tmp_path / "app-data")

    assert os.environ["OPENAI_API_KEY"] == "already-set-key"
    assert os.environ["SCOPELEDGER_PREREVIEW_ENABLED"] == "1"


def test_live_cloudhammer_timeout_env_validation(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SCOPELEDGER_CLOUDHAMMER_TIMEOUT_SECONDS", "not-a-number")

    with pytest.raises(RuntimeError, match="SCOPELEDGER_CLOUDHAMMER_TIMEOUT_SECONDS"):
        LiveCloudHammerPipeline()


def test_production_app_can_load_secret_from_root_env(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.delenv("SCOPELEDGER_WEBAPP_SECRET", raising=False)
    (tmp_path / ".env").write_text("SCOPELEDGER_WEBAPP_SECRET=local-production-secret\n", encoding="utf-8")

    app = create_app(tmp_path / "app-data", production=True)

    assert app.secret_key == "local-production-secret"


def test_production_app_requires_secret(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.chdir(tmp_path)
    monkeypatch.delenv("SCOPELEDGER_WEBAPP_SECRET", raising=False)

    with pytest.raises(RuntimeError, match="SCOPELEDGER_WEBAPP_SECRET"):
        create_app(tmp_path, production=True)


def test_cli_production_serve_requires_secret(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, capsys):
    monkeypatch.chdir(tmp_path)
    monkeypatch.delenv("SCOPELEDGER_WEBAPP_SECRET", raising=False)

    result = cli_main(["serve", str(tmp_path / "app-data"), "--production"])

    captured = capsys.readouterr()
    assert result == 1
    assert "SCOPELEDGER_WEBAPP_SECRET" in captured.out


def test_cli_production_serve_uses_waitress(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    calls = {}

    def fake_serve(app, *, host: str, port: int) -> None:
        calls["host"] = host
        calls["port"] = port
        calls["secret"] = app.secret_key
        calls["production"] = app.config["SCOPELEDGER_PRODUCTION"]

    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")
    monkeypatch.setattr(cli_module, "serve_production_app", fake_serve)

    result = cli_main(["serve", str(tmp_path / "app-data"), "--host", "127.0.0.1", "--port", "5055", "--production"])

    assert result == 0
    assert calls == {
        "host": "127.0.0.1",
        "port": 5055,
        "secret": "test-production-secret",
        "production": True,
    }


def test_cli_production_serve_rejects_public_host(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, capsys):
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")

    result = cli_main(["serve", str(tmp_path / "app-data"), "--host", "0.0.0.0", "--production"])

    captured = capsys.readouterr()
    assert result == 1
    assert "loopback host" in captured.out


def test_cli_serve_rejects_project_workspace_as_app_data_root(tmp_path: Path, capsys):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    input_dir.mkdir()
    WorkspaceStore(workspace_dir).create(input_dir)

    result = cli_main(["serve", str(workspace_dir)])

    captured = capsys.readouterr()
    assert result == 1
    assert "app data root" in captured.out


def test_production_app_sets_release_cookie_and_security_headers(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")

    app = create_app(tmp_path, production=True)
    client = app.test_client()
    response = client.get("/projects")

    assert app.config["SESSION_COOKIE_SECURE"] is True
    assert app.config["SESSION_COOKIE_HTTPONLY"] is True
    assert app.config["SESSION_COOKIE_SAMESITE"] == "Lax"
    assert response.headers["X-Content-Type-Options"] == "nosniff"
    assert response.headers["X-Frame-Options"] == "DENY"
    assert response.headers["Referrer-Policy"] == "same-origin"
    assert response.headers["Cache-Control"] == "no-store"


def test_production_post_requires_csrf(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")

    app = create_app(tmp_path, production=True)
    client = app.test_client()
    response = client.post("/projects", data={"name": "Blocked Project"})

    assert response.status_code == 400
    assert ProjectRegistry(tmp_path).load().projects == []


def test_project_creation_rejects_custom_workspace_paths(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")
    app = create_app(tmp_path, production=True)
    client = app.test_client()
    projects = client.get("/projects")
    token = csrf_token_from(projects)
    assert b"Browse folder" not in projects.data

    response = client.post(
        "/projects",
        data={"name": "Unsafe Path", "workspace_dir": str(tmp_path / "elsewhere"), "csrf_token": token},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Project storage is managed by ScopeLedger" not in response.data
    assert b"flash-stack" not in response.data
    assert ProjectRegistry(tmp_path).load().projects == []


def test_dev_project_creation_also_rejects_custom_workspace_paths(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()

    response = client.post(
        "/projects",
        data={"name": "Unsafe Path", "workspace_dir": str(tmp_path / "elsewhere")},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Project storage is managed by ScopeLedger" not in response.data
    assert b"flash-stack" not in response.data
    assert ProjectRegistry(tmp_path).load().projects == []


def test_project_creation_does_not_render_flash_notifications(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()

    response = client.post("/projects", data={"name": "Fresh Project"}, follow_redirects=True)

    assert response.status_code == 200
    assert b"Created Fresh Project" not in response.data
    assert b"flash-stack" not in response.data
    assert ProjectRegistry(tmp_path).load().projects[0].name == "Fresh Project"


def test_project_creation_uses_managed_app_project_root(tmp_path: Path):
    registry = ProjectRegistry(tmp_path).load()

    project = registry.create_project("Test Revision 2")

    assert Path(project.workspace_dir) == tmp_path / "projects" / "test-revision-2"
    assert Path(project.input_dir) == tmp_path / "projects" / "test-revision-2" / "input"
    assert (tmp_path / "projects" / "test-revision-2" / "workspace.json").exists()


def test_cli_reset_projects_clears_registry_without_deleting_workspace(tmp_path: Path):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    app_data_dir = tmp_path / "app-data"
    input_dir.mkdir()
    store = WorkspaceStore(workspace_dir).create(input_dir)
    registry = ProjectRegistry(app_data_dir).load()
    registry.projects = [
        ProjectRecord(
            id="test-project",
            name="Test Project",
            workspace_dir=str(workspace_dir.resolve()),
            input_dir=store.data.input_dir,
            status="active",
            created_at=store.data.created_at,
        )
    ]
    registry.save()

    result = cli_main(["reset-projects", str(app_data_dir)])

    assert result == 0
    assert (workspace_dir / "workspace.json").exists()
    assert ProjectRegistry(app_data_dir).load().projects == []


def test_cli_reset_projects_rejects_project_workspace_path(tmp_path: Path, capsys):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    input_dir.mkdir()
    WorkspaceStore(workspace_dir).create(input_dir)

    result = cli_main(["reset-projects", str(workspace_dir)])

    captured = capsys.readouterr()
    assert result == 1
    assert "app data root" in captured.out


def test_web_routes_render_without_ai(workspace_copy):
    app = create_app(workspace_copy)
    client = app.test_client()

    assert client.get("/").status_code == 302
    projects = client.get("/projects")
    assert projects.status_code == 200
    assert b"Project archive" in projects.data
    assert b"Workspace folder" not in projects.data
    assert b"Browse folder" not in projects.data
    assert str(workspace_copy).encode("utf-8") not in projects.data
    assert b"Project name" in projects.data
    assert b"Initial package" not in projects.data
    assert b"Manual source path" not in projects.data
    assert b"Input folder" not in projects.data
    assert client.get("/overview").status_code == 200
    assert client.get("/sheets").status_code == 200
    queue = client.get("/changes")
    assert queue.status_code == 200
    assert b"Accept selected" in queue.data
    diagnostics = client.get("/diagnostics")
    assert diagnostics.status_code == 200
    assert b"Ingested PDF files" in diagnostics.data
    settings = client.get("/settings")
    assert settings.status_code == 200
    assert b"Archived" in settings.data


def test_import_revision_sets_root_preserves_child_packages(tmp_path: Path):
    def write_pdf(path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        document = fitz.open()
        document.new_page(width=300, height=200)
        document.save(path)
        document.close()

    source_root = tmp_path / "revision_sets"
    write_pdf(source_root / "Revision #1 - Alpha" / "alpha.pdf")
    write_pdf(source_root / "Revision #2 - Beta" / "beta.pdf")

    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    imported = client.post("/packages/import", data={"source_path": str(source_root), "package_label": ""})
    assert imported.status_code == 302

    input_dir = tmp_path / "projects" / "fresh-project" / "input"
    assert (input_dir / "Revision #1 - Alpha" / "alpha.pdf").exists()
    assert (input_dir / "Revision #2 - Beta" / "beta.pdf").exists()
    assert not (input_dir / "revision_sets" / "Revision #1 - Alpha" / "alpha.pdf").exists()


def test_production_manual_import_allows_configured_root(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    source_root = tmp_path / "revision_sets"
    write_minimal_drawing_pdf(source_root / "Revision #1 - Alpha" / "alpha.pdf")
    write_minimal_drawing_pdf(source_root / "Revision #2 - Beta" / "beta.pdf")
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")
    monkeypatch.setenv("SCOPELEDGER_ALLOWED_IMPORT_ROOTS", str(source_root))

    app = create_app(tmp_path, production=True)
    client = app.test_client()
    token = csrf_token_from(client.get("/projects"))
    created = client.post("/projects", data={"name": "Fresh Project", "csrf_token": token})
    assert created.status_code == 302

    imported = client.post("/packages/import", data={"source_path": str(source_root), "package_label": "", "csrf_token": token})

    assert imported.status_code == 302
    input_dir = tmp_path / "projects" / "fresh-project" / "input"
    assert (input_dir / "Revision #1 - Alpha" / "alpha.pdf").exists()
    assert (input_dir / "Revision #2 - Beta" / "beta.pdf").exists()


def test_production_manual_import_rejects_outside_allowed_root(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    allowed_root = tmp_path / "allowed_revision_sets"
    source_root = tmp_path / "outside_revision_sets"
    allowed_root.mkdir()
    write_minimal_drawing_pdf(source_root / "Revision #1 - Outside" / "outside.pdf")
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")
    monkeypatch.setenv("SCOPELEDGER_ALLOWED_IMPORT_ROOTS", str(allowed_root))

    app = create_app(tmp_path, production=True)
    client = app.test_client()
    token = csrf_token_from(client.get("/projects"))
    created = client.post("/projects", data={"name": "Fresh Project", "csrf_token": token})
    assert created.status_code == 302

    imported = client.post(
        "/packages/import",
        data={"source_path": str(source_root), "package_label": "", "csrf_token": token},
        follow_redirects=True,
    )

    assert imported.status_code == 200
    assert b"outside the configured production import allowlist" not in imported.data
    assert b"flash-stack" not in imported.data
    input_dir = tmp_path / "projects" / "fresh-project" / "input"
    assert not (input_dir / "Revision #1 - Outside" / "outside.pdf").exists()


def test_chunked_browser_import_reconstructs_pdf_package(tmp_path: Path):
    payload = b"%PDF-1.7\n" + (b"0" * (8 * 1024 * 1024)) + b"\n%%EOF"
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    init = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "package_label": "Chunked Package",
            "files": [{"name": "large_package.pdf", "relative_path": "large_package.pdf", "size": len(payload)}],
        },
    )
    assert init.status_code == 200
    upload_id = init.get_json()["upload_id"]

    first_chunk = payload[: 8 * 1024 * 1024]
    second_chunk = payload[8 * 1024 * 1024 :]
    for index, chunk in enumerate([first_chunk, second_chunk]):
        response = client.post(
            "/uploads/chunked/chunk",
            data={
                "upload_id": upload_id,
                "file_index": "0",
                "chunk_index": str(index),
                "chunk": (io.BytesIO(chunk), f"chunk-{index}.part"),
            },
            content_type="multipart/form-data",
        )
        assert response.status_code == 200

    complete = client.post("/uploads/chunked/complete", json={"upload_id": upload_id})

    assert complete.status_code == 200
    input_dir = tmp_path / "projects" / "fresh-project" / "input"
    assert (input_dir / "Chunked Package" / "large_package.pdf").read_bytes() == payload
    assert not (tmp_path / "projects" / "fresh-project" / ".chunked_uploads" / upload_id).exists()


def test_chunked_upload_rejects_non_pdf(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    response = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "files": [{"name": "notes.txt", "relative_path": "notes.txt", "size": 10}],
        },
    )

    assert response.status_code == 400
    assert "Only PDF files" in response.get_json()["error"]


def test_production_chunked_upload_requires_csrf(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")
    app = create_app(tmp_path, production=True)
    client = app.test_client()
    token = csrf_token_from(client.get("/projects"))
    created = client.post("/projects", data={"name": "Fresh Project", "csrf_token": token})
    assert created.status_code == 302

    response = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "files": [{"name": "drawing.pdf", "relative_path": "drawing.pdf", "size": 12}],
        },
    )

    assert response.status_code == 400
    assert "form token" in response.get_json()["error"]


def test_chunked_upload_respects_configured_size_limit(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("SCOPELEDGER_MAX_UPLOAD_BYTES", str(20 * 1024 * 1024))
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    response = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "files": [{"name": "huge.pdf", "relative_path": "huge.pdf", "size": 21 * 1024 * 1024}],
        },
    )

    assert response.status_code == 413
    assert "upload limit" in response.get_json()["error"]


def test_production_workspace_folder_dialog_is_disabled(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("SCOPELEDGER_WEBAPP_SECRET", "test-production-secret")
    app = create_app(tmp_path, production=True)
    client = app.test_client()

    response = client.get("/system/dialog/workspace-folder")

    assert response.status_code == 404
    assert "disabled" in response.get_json()["error"]


def test_chunked_upload_abort_removes_temp_session(tmp_path: Path):
    payload = b"%PDF-1.7\n%%EOF"
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    init = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "files": [{"name": "cancel_me.pdf", "relative_path": "cancel_me.pdf", "size": len(payload)}],
        },
    )
    assert init.status_code == 200
    upload_id = init.get_json()["upload_id"]
    chunk = client.post(
        "/uploads/chunked/chunk",
        data={
            "upload_id": upload_id,
            "file_index": "0",
            "chunk_index": "0",
            "chunk": (io.BytesIO(payload), "cancel.part"),
        },
        content_type="multipart/form-data",
    )
    assert chunk.status_code == 200
    upload_dir = tmp_path / "projects" / "fresh-project" / ".chunked_uploads" / upload_id
    assert upload_dir.exists()

    abort = client.post("/uploads/chunked/abort", json={"upload_id": upload_id})

    assert abort.status_code == 200
    assert abort.get_json()["deleted"] is True
    assert not upload_dir.exists()


def test_chunked_upload_init_cleans_stale_temp_sessions(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302
    upload_root = tmp_path / "projects" / "fresh-project" / ".chunked_uploads"
    stale_dir = upload_root / ("a" * 32)
    current_dir = upload_root / ("b" * 32)
    stale_dir.mkdir(parents=True)
    current_dir.mkdir(parents=True)
    (stale_dir / "metadata.json").write_text("{}", encoding="utf-8")
    (current_dir / "metadata.json").write_text("{}", encoding="utf-8")
    old_time = time.time() - (7 * 60 * 60)
    os.utime(stale_dir, (old_time, old_time))

    init = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "files": [{"name": "fresh.pdf", "relative_path": "fresh.pdf", "size": 12}],
        },
    )

    assert init.status_code == 200
    assert not stale_dir.exists()
    assert current_dir.exists()


def test_chunked_browser_append_reconstructs_pdf(tmp_path: Path):
    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    revision_source_dir = input_dir / "Revision #1 - Base"
    write_minimal_drawing_pdf(revision_source_dir / "base.pdf", scope_text="BASE PACKAGE")
    store = WorkspaceStore(workspace_dir).create(input_dir)
    revision_set = RevisionSet(
        id="rev-1",
        label="Revision #1 - Base",
        source_dir=str(revision_source_dir),
        set_number=1,
        set_date="05/09/2026",
    )
    store.data.revision_sets = [revision_set]
    store.save()
    register_workspace_project(workspace_dir)

    app = create_app(workspace_dir)
    client = app.test_client()
    appended_source = tmp_path / "appendix.pdf"
    write_minimal_drawing_pdf(appended_source, scope_text="APPENDED VALID PDF")
    payload = appended_source.read_bytes()

    init = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "append_file",
            "revision_set_id": revision_set.id,
            "files": [{"name": "appendix.pdf", "relative_path": "appendix.pdf", "size": len(payload)}],
        },
    )
    assert init.status_code == 200
    upload_id = init.get_json()["upload_id"]
    chunk = client.post(
        "/uploads/chunked/chunk",
        data={
            "upload_id": upload_id,
            "file_index": "0",
            "chunk_index": "0",
            "chunk": (io.BytesIO(payload), "appendix.part"),
        },
        content_type="multipart/form-data",
    )
    assert chunk.status_code == 200

    complete = client.post("/uploads/chunked/complete", json={"upload_id": upload_id})

    assert complete.status_code == 200
    destination = WorkspaceStore(workspace_dir).load().resolve_path(revision_set.source_dir) / "appendix.pdf"
    assert destination.read_bytes() == payload


def test_manual_import_rejects_folder_without_pdfs(tmp_path: Path):
    source_root = tmp_path / "not-a-package"
    source_root.mkdir()
    (source_root / "notes.txt").write_text("not a drawing", encoding="utf-8")

    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    imported = client.post("/packages/import", data={"source_path": str(source_root), "package_label": ""})

    assert imported.status_code == 302
    input_dir = tmp_path / "projects" / "fresh-project" / "input"
    assert not (input_dir / "not-a-package" / "notes.txt").exists()


def test_populate_workspace_runs_cloudhammer_manifest_from_ui(tmp_path: Path):
    class FakeCloudHammerRunner:
        name = "fake_cloudhammer_live"

        def run(self, *, input_dir: Path, workspace_dir: Path) -> CloudHammerRunResult:
            pdf_path = next(input_dir.rglob("*.pdf"))
            run_dir = workspace_dir / "outputs" / "cloudhammer_live" / "fake"
            run_dir.mkdir(parents=True, exist_ok=True)
            crop_path = run_dir / "crop.png"
            crop_path.write_bytes(b"png")
            pages_manifest = run_dir / "pages_manifest.jsonl"
            pages_manifest.write_text(
                json.dumps(
                    {
                        "page_kind": "drawing",
                        "pdf_path": str(pdf_path),
                        "pdf_stem": pdf_path.stem,
                        "page_index": 0,
                        "page_number": 1,
                        "render_path": str(run_dir / "render.png"),
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            candidate_manifest = run_dir / "whole_cloud_candidates_manifest.jsonl"
            candidate_manifest.write_text(
                json.dumps(
                    {
                        "candidate_id": "fake-live-1",
                        "pdf_path": str(pdf_path),
                        "page_number": 1,
                        "bbox_page_xywh": [70, 70, 120, 80],
                        "whole_cloud_confidence": 0.91,
                        "policy_bucket": "auto_deliverable_candidate",
                        "crop_image_path": str(crop_path),
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            return CloudHammerRunResult(
                run_dir=run_dir,
                pages_manifest=pages_manifest,
                candidate_manifest=candidate_manifest,
                page_count=1,
                candidate_count=1,
            )

    app = create_app(tmp_path, cloudhammer_runner=FakeCloudHammerRunner())
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    input_dir = tmp_path / "projects" / "fresh-project" / "input"
    write_minimal_drawing_pdf(input_dir / "Revision #1 - Test" / "drawing.pdf")

    populated = client.post("/workspace/populate")

    assert populated.status_code == 302
    store = WorkspaceStore(tmp_path / "projects" / "fresh-project").load()
    assert len(store.data.clouds) == 1
    assert any(item.provenance.get("extraction_method") == "cloudhammer_manifest" for item in store.data.change_items)
    assert store.data.populate_status["cloudhammer_page_count"] == 1
    assert store.data.populate_status["cloudhammer_candidate_count"] == 1


def test_populate_status_endpoint_reports_staged_and_live_artifacts(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    workspace_dir = tmp_path / "projects" / "fresh-project"
    input_dir = workspace_dir / "input"
    write_minimal_drawing_pdf(input_dir / "Revision #1 - Test" / "drawing.pdf")
    run_dir = workspace_dir / "outputs" / "cloudhammer_live" / "run_test"
    run_dir.mkdir(parents=True)
    (run_dir / "pages_manifest.jsonl").write_text(json.dumps({"page_kind": "drawing"}) + "\n", encoding="utf-8")
    candidates_dir = run_dir / "whole_cloud_candidates"
    candidates_dir.mkdir()
    (candidates_dir / "whole_cloud_candidates_manifest.jsonl").write_text(json.dumps({"candidate_id": "c1"}) + "\n", encoding="utf-8")
    store = WorkspaceStore(workspace_dir).load()
    store.update_populate_status(
        state="running",
        stage="drawing_analysis",
        message="Analyzing staged drawing packages.",
        pre_review_total_count=12,
        pre_review_2_count=7,
        pre_review_failed_count=1,
        pre_review_cache_hits=3,
    )

    response = client.get("/workspace/populate/status")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["state"] == "running"
    assert payload["staged_pdf_count"] == 1
    assert payload["staged_package_count"] == 1
    assert payload["live_artifact_count"] == 2
    assert payload["inferred_cloudhammer_page_count"] == 1
    assert payload["inferred_cloudhammer_candidate_count"] == 1
    assert payload["pre_review_total_count"] == 12
    assert payload["pre_review_2_count"] == 7
    assert payload["pre_review_failed_count"] == 1
    assert payload["pre_review_cache_hits"] == 3


def test_dashboard_exposes_populate_polling_hooks(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()
    created = client.post("/projects", data={"name": "Fresh Project"})
    assert created.status_code == 302

    response = client.get("/overview")

    assert response.status_code == 200
    assert b"js-populate-form" in response.data
    assert b'data-status-url="/workspace/populate/status"' in response.data
    assert b'data-populate-status-url="/workspace/populate/status"' in response.data
    assert b"Staged PDFs" in response.data
    assert b"Live artifacts" in response.data
    assert b"Pre Review" in response.data
    assert b"GPT" not in response.data
    assert b"CloudHammer" not in response.data
    assert b"training" not in response.data
    assert b"eval" not in response.data
    assert b"labeling" not in response.data


def test_review_routes_reject_invalid_status_and_external_redirect(workspace_copy):
    app = create_app(workspace_copy)
    client = app.test_client()
    store = WorkspaceStore(workspace_copy).load()
    item = next(item for item in store.data.change_items if item.status == "pending")

    response = client.post(
        f"/changes/{item.id}/review",
        data={
            "status_override": "surprise",
            "reviewer_text": item.raw_text,
            "reviewer_notes": "",
            "redirect_to": "https://example.com/owned",
        },
    )

    assert response.status_code == 302
    assert response.headers["Location"].endswith(f"/changes/{item.id}")
    refreshed = WorkspaceStore(workspace_copy).load()
    assert refreshed.get_change_item(item.id).status == "pending"

    bulk = client.post(
        "/changes/bulk-review",
        data={"change_ids": [item.id], "status": "approved", "redirect_to": "https://example.com/owned"},
    )
    assert bulk.status_code == 302
    assert bulk.headers["Location"].endswith("/changes")


def test_project_asset_route_blocks_non_generated_files(workspace_copy):
    app = create_app(workspace_copy)
    client = app.test_client()

    response = client.get("/project-assets/docs/CURRENT_STATE.md")

    assert response.status_code == 404


def test_change_detail_uses_previous_current_comparison_asset(tmp_path: Path):
    def write_color_pdf(path: Path, fill: tuple[float, float, float]) -> None:
        document = fitz.open()
        page = document.new_page(width=300, height=200)
        page.draw_rect(page.rect, color=fill, fill=fill)
        document.save(path)
        document.close()

    input_dir = tmp_path / "input"
    workspace_dir = tmp_path / "workspace"
    input_dir.mkdir()
    previous_pdf = input_dir / "previous.pdf"
    current_pdf = input_dir / "current.pdf"
    write_color_pdf(previous_pdf, (1, 0, 0))
    write_color_pdf(current_pdf, (0, 0, 1))
    store = WorkspaceStore(workspace_dir).create(input_dir)
    store.data.revision_sets = [
        RevisionSet(id="rev-1", label="Revision #1", source_dir=str(input_dir), set_number=1, set_date="04/01/2026"),
        RevisionSet(id="rev-2", label="Revision #2", source_dir=str(input_dir), set_number=2, set_date="04/15/2026"),
    ]
    previous_sheet = SheetVersion(
        id="sheet-prev",
        revision_set_id="rev-1",
        source_pdf=str(previous_pdf),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Plan",
        issue_date="04/01/2026",
        width=300,
        height=200,
        status="superseded",
    )
    current_sheet = SheetVersion(
        id="sheet-current",
        revision_set_id="rev-2",
        source_pdf=str(current_pdf),
        page_number=1,
        sheet_id="AE101",
        sheet_title="Plan",
        issue_date="04/15/2026",
        width=300,
        height=200,
        status="active",
    )
    cloud = CloudCandidate(
        id="cloud-1",
        sheet_version_id=current_sheet.id,
        bbox=[80, 60, 80, 60],
        image_path="",
        page_image_path="",
        confidence=0.9,
        extraction_method="cloudhammer_manifest",
        nearby_text="Cloud Only",
        detail_ref=None,
    )
    store.data.sheets = [previous_sheet, current_sheet]
    store.data.clouds = [cloud]
    store.data.change_items = [
        ChangeItem(
            id="change-1",
            sheet_version_id=current_sheet.id,
            cloud_candidate_id=cloud.id,
            sheet_id=current_sheet.sheet_id,
            detail_ref=None,
            raw_text="Cloud Only",
            normalized_text="cloud only",
            provenance={"source": "visual-region", "extraction_method": "cloudhammer_manifest"},
        )
    ]
    store.save()
    register_workspace_project(workspace_dir)

    app = create_app(workspace_dir)
    client = app.test_client()
    response = client.get("/changes/change-1")
    assert response.status_code == 200
    assert b"/workspace-assets/pre-review/change-1.png" in response.data

    image_response = client.get("/workspace-assets/cloud-comparisons/cloud-1.png")
    assert image_response.status_code == 200
    assert image_response.data[:8] == b"\x89PNG\r\n\x1a\n"


def test_dashboard_shows_pricing_readiness_panel(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    store.update_populate_status(
        state="done",
        stage="complete",
        message="Workspace is populated and ready for review.",
        package_count=2,
        document_count=2,
        sheet_count=len(store.data.sheets),
        cloud_count=len(store.data.clouds),
        change_item_count=len(store.data.change_items),
        cache_hits=2,
    )
    app = create_app(workspace_copy)
    client = app.test_client()
    response = client.get("/overview")
    assert response.status_code == 200
    body = response.data
    assert b"Revision packages" in body
    assert b"Export readiness" in body
    assert b"Pending review" in body
    assert b"Accepted" in body
    assert b"Needs check" in body
    assert b"Project Files and Packages" in body
    assert b"Choose PDF(s)" in body
    assert b"Choose folder" in body
    assert b"Manual server path" in body
    assert b'name="package_files"' in body
    assert b'name="append_files"' in body
    assert b"Populate Workspace" in body
    assert b"Populate status" in body
    assert b"Workspace is populated and ready for review." in body
    assert b"Cache hits" in body


def test_dashboard_and_export_link_generated_artifacts(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    Exporter(store).export(force_attention=True)
    packet = build_review_packet(store)

    app = create_app(workspace_copy)
    client = app.test_client()

    dashboard = client.get("/overview")
    assert dashboard.status_code == 200
    assert b"Open Workbook" not in dashboard.data
    assert b"Open Review Packet" not in dashboard.data

    export = client.get("/export")
    assert export.status_code == 200
    assert b"Open Workbook" in export.data
    assert b'href="/outputs/revision_changelog.xlsx"' in export.data
    assert b"Re-generate workbook" in export.data
    assert b"Open Review Packet" in export.data
    assert b'href="/outputs/revision_changelog_review_packet.html"' in export.data
    assert b"Refresh Review Packet" in export.data
    assert b"Open Google Drive Folder" in export.data

    workbook_response = client.get("/outputs/revision_changelog.xlsx")
    assert workbook_response.status_code == 200
    assert workbook_response.data[:2] == b"PK"

    packet_response = client.get(f"/outputs/{packet.html_path.name}")
    assert packet_response.status_code == 200
    assert b"ScopeLedger Review Packet" in packet_response.data


def test_conformed_page_lists_revised_sheets_by_default(workspace_copy):
    app = create_app(workspace_copy)
    client = app.test_client()

    response = client.get("/conformed")
    assert response.status_code == 200
    body = response.data
    assert b"Latest Set" in body
    assert b"Revised only" in body
    assert b"Revised" in body

    response_all = client.get("/conformed?show=all")
    assert response_all.status_code == 200
    assert response_all.data.count(b"conformed-card") >= response.data.count(b"conformed-card")


def test_navbar_includes_conformed_link(workspace_copy):
    app = create_app(workspace_copy)
    client = app.test_client()
    response = client.get("/projects")
    assert b'href="/conformed"' in response.data
    assert b"Latest Set" in response.data


def test_bulk_review_and_next_navigation(workspace_copy):
    app = create_app(workspace_copy, bulk_review_manager=BulkReviewJobManager(run_async=False))
    client = app.test_client()
    store = WorkspaceStore(workspace_copy).load()
    pending_items = [item for item in store.data.change_items if item.status == "pending"][:2]

    response = client.post(
        "/changes/bulk-review",
        data={"change_ids": [item.id for item in pending_items], "status": "approved", "redirect_to": "/changes?status=pending"},
    )
    assert response.status_code == 302

    refreshed = WorkspaceStore(workspace_copy).load()
    assert all(refreshed.get_change_item(item.id).status == "approved" for item in pending_items)

    remaining_pending = [item for item in refreshed.data.change_items if item.status == "pending"][:2]
    detail = client.get(f"/changes/{remaining_pending[0].id}?queue=pending")
    assert detail.status_code == 200
    assert b"Save + Next" not in detail.data

    advance = client.post(
        f"/changes/{remaining_pending[0].id}/review",
        data={
            "status": "approved",
            "reviewer_text": "Fast lane approval",
            "reviewer_notes": "",
            "queue_status": "pending",
            "search_query": "",
            "next_change_id": remaining_pending[1].id,
            "advance": "next",
        },
    )
    assert advance.status_code == 302
    assert advance.headers["Location"].endswith(f"/changes/{remaining_pending[1].id}?queue=pending&q=")

    reloaded = WorkspaceStore(workspace_copy).load()
    assert reloaded.get_change_item(remaining_pending[0].id).reviewer_text == "Fast lane approval"
    assert reloaded.get_change_item(remaining_pending[0].id).status == "approved"


def test_rescan_reuses_cache_and_preserves_review_state(workspace_copy):
    store = WorkspaceStore(workspace_copy).load()
    change = next(item for item in store.data.change_items if item.status == "pending")
    store.update_change_item(change.id, status="approved", reviewer_text="Preserved after rescan")
    render_path = Path(store.data.sheets[0].render_path)
    original_mtime = render_path.stat().st_mtime_ns
    input_dir = Path(store.data.input_dir)

    scanner = RevisionScanner(input_dir, workspace_copy)
    scanner.scan()

    reloaded = WorkspaceStore(workspace_copy).load()
    rescanned = reloaded.get_change_item(change.id)
    assert scanner.cache_hits == len(reloaded.data.documents)
    assert rescanned.status == "approved"
    assert rescanned.reviewer_text == "Preserved after rescan"
    assert render_path.stat().st_mtime_ns == original_mtime
    assert len(reloaded.data.scan_cache.get("documents", {})) == len(reloaded.data.documents)


def test_verify_endpoint_is_archived(workspace_copy):
    app = create_app(workspace_copy)
    client = app.test_client()
    store = WorkspaceStore(workspace_copy).load()
    change = store.data.change_items[0]

    response = client.post(f"/changes/{change.id}/verify")
    assert response.status_code == 400
    payload = response.get_json()
    assert "disabled" in payload["error"].lower() or "archived" in payload["error"].lower()
