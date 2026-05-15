from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook

from backend.deliverables.review_packet import build_review_packet
from backend.projects import ProjectRegistry
from backend.review_queue import visible_change_items
from backend.workspace import WorkspaceStore
from webapp.app import create_app

from tests.smoke_helpers import build_smoke_workspace, mark_package_runs_complete, write_smoke_pdf


def test_smoke_empty_project_shell_renders(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()

    projects = client.get("/projects")
    changes = client.get("/changes")

    assert projects.status_code == 200
    assert b"No active projects." in projects.data
    assert b"Project name" in projects.data
    assert changes.status_code == 200
    assert b"Review Changes" in changes.data


def test_smoke_active_handoff_routes_render(tmp_path: Path):
    smoke = build_smoke_workspace(tmp_path)
    app = create_app(smoke.app_data_dir)
    client = app.test_client()

    for path, marker in [
        ("/projects", b"Smoke Project"),
        ("/overview", b"Revision packages"),
        ("/sheets", b"PL505"),
        ("/changes", b"Review Changes"),
        ("/changes/change-r2?queue=pending&package_scope=newest", b"Mark as legend"),
        ("/sheets/sheet-r2", b'data-coordinate-width="400"'),
        ("/export", b"Export Workbook"),
    ]:
        response = client.get(path)
        assert response.status_code == 200, path
        assert marker in response.data, path

    assert b"Open Google Drive Folder" not in client.get("/export").data


def test_smoke_chunked_upload_import_stages_package(tmp_path: Path):
    app = create_app(tmp_path)
    client = app.test_client()
    assert client.post("/projects", data={"name": "Upload Smoke"}).status_code == 302
    source_pdf = tmp_path / "upload.pdf"
    write_smoke_pdf(source_pdf)
    payload = source_pdf.read_bytes()

    init = client.post(
        "/uploads/chunked/init",
        json={
            "purpose": "import_package",
            "package_label": "Chunked Smoke",
            "revision_number": 7,
            "files": [{"name": "upload.pdf", "relative_path": "upload.pdf", "size": len(payload)}],
        },
    )
    assert init.status_code == 200
    upload_id = init.get_json()["upload_id"]
    chunk = client.post(
        "/uploads/chunked/chunk",
        data={
            "upload_id": upload_id,
            "file_index": "0",
            "chunk_index": "0",
            "chunk": (io.BytesIO(payload), "upload.part"),
        },
        content_type="multipart/form-data",
    )
    assert chunk.status_code == 200

    complete = client.post("/uploads/chunked/complete", json={"upload_id": upload_id})

    assert complete.status_code == 200
    store = WorkspaceStore(tmp_path / "projects" / "upload-smoke").load()
    assert store.data.staged_packages[0].label == "Chunked Smoke"
    assert store.data.staged_packages[0].revision_number == 7
    assert (Path(store.data.input_dir) / "Chunked Smoke" / "upload.pdf").exists()


def test_smoke_populate_status_reports_package_history_rows(tmp_path: Path):
    smoke = build_smoke_workspace(tmp_path)
    app = create_app(smoke.app_data_dir)
    mark_package_runs_complete(smoke.store, app.config["CLOUDHAMMER_RUNNER"])
    client = app.test_client()

    response = client.get("/workspace/populate/status")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["total_package_count"] == 2
    assert payload["reusable_package_count"] == 2
    assert payload["dirty_package_count"] == 0
    assert {row["status"] for row in payload["package_run_rows"]} == {"reused"}
    assert {row["revision_number"] for row in payload["package_run_rows"]} == {1, 2}


def test_smoke_review_package_filters_and_detail_controls(tmp_path: Path):
    smoke = build_smoke_workspace(tmp_path)
    app = create_app(smoke.app_data_dir)
    client = app.test_client()

    newest = client.get("/changes?package_scope=newest")
    revision_1 = client.get("/changes?package_scope=package&package_id=pkg-r1")
    detail = client.get("/changes/change-r2?queue=pending&package_scope=newest")

    assert newest.status_code == 200
    assert b"Smoke scope revision 2" in newest.data
    assert b"Smoke scope revision 1" not in newest.data
    assert revision_1.status_code == 200
    assert b"Smoke scope revision 1" in revision_1.data
    assert b"Smoke scope revision 2" not in revision_1.data
    assert detail.status_code == 200
    assert b'name="package_scope" value="newest"' in detail.data
    assert b"Mark as legend" in detail.data


def test_smoke_manual_mark_as_legend_hides_item_from_queue(tmp_path: Path):
    smoke = build_smoke_workspace(tmp_path)
    app = create_app(smoke.app_data_dir)
    client = app.test_client()

    response = client.post(
        "/changes/change-r1/accept-legend",
        data={"queue_status": "pending", "search_query": "", "attention_only": "0", "package_scope": "all"},
    )

    assert response.status_code == 302
    loaded = WorkspaceStore(smoke.project.workspace_dir).load()
    visible_ids = [item.id for item in visible_change_items(loaded.data.change_items)]
    assert visible_ids == ["change-r2"]
    queue = client.get("/changes")
    assert b"Smoke scope revision 1" not in queue.data
    assert b"Smoke scope revision 2" in queue.data


def test_smoke_sheet_overlay_coordinates_render(tmp_path: Path):
    smoke = build_smoke_workspace(tmp_path)
    app = create_app(smoke.app_data_dir)
    client = app.test_client()

    response = client.get("/sheets/sheet-r1")

    assert response.status_code == 200
    assert b'data-coordinate-width="400"' in response.data
    assert b'data-coordinate-height="240"' in response.data
    assert b'data-x="80.0"' in response.data
    assert b'data-w="120.0"' in response.data


def test_smoke_export_outputs_workbook_and_review_packet(tmp_path: Path):
    smoke = build_smoke_workspace(tmp_path)
    store = WorkspaceStore(smoke.project.workspace_dir).load()
    store.update_change_item("change-r1", status="approved", reviewer_text="Approved smoke scope")
    store = WorkspaceStore(smoke.project.workspace_dir).load()
    store.update_change_item("change-r2", status="approved", reviewer_text="Approved smoke scope revision 2")
    app = create_app(smoke.app_data_dir)
    client = app.test_client()

    response = client.post("/export/run")

    assert response.status_code == 302
    packet_response = client.post("/review-packet/run")
    assert packet_response.status_code == 302
    export = client.get("/export")
    assert b"Open Workbook" in export.data
    assert b"Open Review Packet" in export.data
    assert b"Open Google Drive Folder" not in export.data
    workbook_path = Path(smoke.project.workspace_dir) / "outputs" / "revision_changelog.xlsx"
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    assert "\n" in workbook["Sheet1"].cell(row=2, column=3).value
    packet = build_review_packet(WorkspaceStore(smoke.project.workspace_dir).load())
    assert packet.item_count == 2
    assert packet.html_path.exists()


def test_smoke_project_registry_selects_active_project(tmp_path: Path):
    registry = ProjectRegistry(tmp_path).load()
    first = registry.create_project("First Smoke")
    second = registry.create_project("Second Smoke")
    app = create_app(tmp_path)
    client = app.test_client()

    response = client.post(f"/projects/{second.id}/select", follow_redirects=True)

    assert response.status_code == 200
    assert b"Second Smoke" in response.data
    assert b"First Smoke" not in response.data
